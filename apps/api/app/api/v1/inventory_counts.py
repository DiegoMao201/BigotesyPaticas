"""Sistema de conteo físico de inventario.

Flujo completo:
  1. POST /inventory-counts          → Crear sesión de conteo (snapshot de stock actual)
  2. GET  /inventory-counts/{id}/template → Descargar plantilla Excel para contar
  3. POST /inventory-counts/{id}/upload   → Subir Excel diligenciado (preview diferencias)
  4. POST /inventory-counts/{id}/apply    → Aplicar ajustes a stock (genera StockMovements)
  5. GET  /inventory-counts               → Historial de sesiones con métricas
  6. GET  /inventory-counts/{id}          → Detalle completo sesión + items + diferencias
  7. DELETE /inventory-counts/{id}        → Eliminar sesión (solo si no aplicada)
"""
from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime
from decimal import Decimal
from typing import Literal

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field
from sqlalchemy import func, or_, select

from app.deps import CurrentUser, DBSession, require_permission
from app.models.catalog import Product
from app.models.inventory import (
    CountItem,
    CountSession,
    Stock,
    StockLocation,
    StockMovement,
)

router = APIRouter(prefix="/inventory-counts", tags=["inventory-counts"])


# ─── Schemas ─────────────────────────────────────────────────────────────────

class CountSessionCreate(BaseModel):
    name: str = Field(min_length=2, max_length=200)
    notes: str | None = None


class CountSessionOut(BaseModel):
    id: uuid.UUID
    name: str
    status: str
    notes: str | None
    total_products_counted: int
    total_with_difference: int
    total_positive_delta: int
    total_negative_delta: int
    total_value_impact: float
    applied_at: datetime | None
    applied_by: str | None
    created_by: str | None
    created_at: datetime
    updated_at: datetime
    items_count: int = 0

    class Config:
        from_attributes = True


class CountItemOut(BaseModel):
    id: uuid.UUID
    product_id: uuid.UUID
    sku: str
    product_name: str
    category_name: str | None
    unit_cost: float
    system_qty: int
    counted_qty: int | None
    delta: int | None
    value_impact: float | None
    notes: str | None

    class Config:
        from_attributes = True


class CountSessionDetail(CountSessionOut):
    items: list[CountItemOut]


class UpdateCountItems(BaseModel):
    """Actualiza conteos de items (puede ser parcial)."""
    items: list[dict]  # [{product_id: str, counted_qty: int, notes?: str}]


class UploadPreviewRow(BaseModel):
    sku: str
    product_name: str
    category_name: str | None
    system_qty: int
    counted_qty: int
    delta: int
    value_impact: float
    unit_cost: float
    status: Literal["ok", "surplus", "shortage", "not_found"]


class UploadPreview(BaseModel):
    matched: int
    not_found: int
    with_difference: int
    total_value_impact: float
    rows: list[UploadPreviewRow]


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _session_or_404(session: CountSession | None) -> CountSession:
    if session is None:
        raise HTTPException(404, "Sesión de conteo no encontrada")
    return session


def _require_status(session: CountSession, *allowed: str) -> None:
    if session.status not in allowed:
        raise HTTPException(
            409,
            f"La sesión está en estado '{session.status}'. "
            f"Esta operación requiere: {', '.join(allowed)}",
        )


async def _get_default_location(db) -> StockLocation:
    loc = (
        await db.execute(
            select(StockLocation).where(StockLocation.is_default == 1).limit(1)
        )
    ).scalar_one_or_none()
    if loc is None:
        raise HTTPException(400, "No hay bodega por defecto configurada")
    return loc


# ─── Endpoints ───────────────────────────────────────────────────────────────

@router.post(
    "",
    response_model=CountSessionOut,
    status_code=201,
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def create_count_session(
    payload: CountSessionCreate,
    db: DBSession,
    user: CurrentUser,
):
    """Crea una sesión de conteo con snapshot del stock actual de todos los productos activos."""
    # Snapshot stock + products
    stock_sub = (
        select(
            Stock.product_id,
            func.coalesce(func.sum(Stock.quantity), 0).label("qty"),
        )
        .group_by(Stock.product_id)
        .subquery()
    )

    # Note: Product has lazy="joined" on .category, so we do NOT add an explicit
    # Category join here — that would create a duplicate join and a 500 error.
    rows = (
        await db.execute(
            select(Product, stock_sub.c.qty)
            .outerjoin(stock_sub, stock_sub.c.product_id == Product.id)
            .where(Product.deleted_at.is_(None))
            .where(Product.is_active == True)  # noqa: E712
            .order_by(Product.name)
        )
    ).all()

    session = CountSession(
        name=payload.name,
        status="draft",
        notes=payload.notes,
        created_by=user.email,
    )
    db.add(session)
    await db.flush()

    count_items = [
        CountItem(
            session_id=session.id,
            product_id=p.id,
            sku=p.sku,
            product_name=p.name,
            category_name=p.category.name if p.category else None,
            unit_cost=Decimal(str(p.cost or 0)),
            system_qty=int(qty or 0),
        )
        for p, qty in rows
    ]
    db.add_all(count_items)
    await db.commit()
    await db.refresh(session)

    return CountSessionOut(
        id=session.id,
        name=session.name,
        status=session.status,
        notes=session.notes,
        total_products_counted=session.total_products_counted,
        total_with_difference=session.total_with_difference,
        total_positive_delta=session.total_positive_delta,
        total_negative_delta=session.total_negative_delta,
        total_value_impact=float(session.total_value_impact),
        applied_at=session.applied_at,
        applied_by=session.applied_by,
        created_by=session.created_by,
        created_at=session.created_at,
        updated_at=session.updated_at,
        items_count=len(count_items),
    )


@router.get(
    "",
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def list_count_sessions(
    db: DBSession,
    status: str | None = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
):
    """Lista historial de sesiones de conteo con métricas."""
    stmt = select(CountSession)
    if status:
        stmt = stmt.where(CountSession.status == status)
    stmt = stmt.order_by(CountSession.created_at.desc())

    total = (
        await db.execute(select(func.count()).select_from(stmt.subquery()))
    ).scalar_one()

    sessions = (
        await db.execute(stmt.offset((page - 1) * page_size).limit(page_size))
    ).scalars().all()

    # Count items per session
    items_count_map: dict[uuid.UUID, int] = {}
    if sessions:
        ids = [s.id for s in sessions]
        counts = (
            await db.execute(
                select(CountItem.session_id, func.count(CountItem.id).label("cnt"))
                .where(CountItem.session_id.in_(ids))
                .group_by(CountItem.session_id)
            )
        ).all()
        items_count_map = {row.session_id: row.cnt for row in counts}

    return {
        "items": [
            CountSessionOut(
                id=s.id,
                name=s.name,
                status=s.status,
                notes=s.notes,
                total_products_counted=s.total_products_counted,
                total_with_difference=s.total_with_difference,
                total_positive_delta=s.total_positive_delta,
                total_negative_delta=s.total_negative_delta,
                total_value_impact=float(s.total_value_impact),
                applied_at=s.applied_at,
                applied_by=s.applied_by,
                created_by=s.created_by,
                created_at=s.created_at,
                updated_at=s.updated_at,
                items_count=items_count_map.get(s.id, 0),
            )
            for s in sessions
        ],
        "total": total,
        "page": page,
        "page_size": page_size,
    }


@router.get(
    "/{session_id}",
    response_model=CountSessionDetail,
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def get_count_session(session_id: uuid.UUID, db: DBSession):
    session = _session_or_404(
        (
            await db.execute(
                select(CountSession).where(CountSession.id == session_id)
            )
        ).scalar_one_or_none()
    )
    return CountSessionDetail(
        id=session.id,
        name=session.name,
        status=session.status,
        notes=session.notes,
        total_products_counted=session.total_products_counted,
        total_with_difference=session.total_with_difference,
        total_positive_delta=session.total_positive_delta,
        total_negative_delta=session.total_negative_delta,
        total_value_impact=float(session.total_value_impact),
        applied_at=session.applied_at,
        applied_by=session.applied_by,
        created_by=session.created_by,
        created_at=session.created_at,
        updated_at=session.updated_at,
        items_count=len(session.items),
        items=[
            CountItemOut(
                id=it.id,
                product_id=it.product_id,
                sku=it.sku,
                product_name=it.product_name,
                category_name=it.category_name,
                unit_cost=float(it.unit_cost),
                system_qty=it.system_qty,
                counted_qty=it.counted_qty,
                delta=it.delta,
                value_impact=float(it.value_impact) if it.value_impact is not None else None,
                notes=it.notes,
            )
            for it in session.items
        ],
    )


@router.get(
    "/{session_id}/template",
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def download_template(session_id: uuid.UUID, db: DBSession):
    """Descarga plantilla Excel (.xlsx) con todos los productos para diligenciar conteo."""
    try:
        import openpyxl
        from openpyxl.styles import (
            Alignment, Border, Font, PatternFill, Side
        )
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado")

    session = _session_or_404(
        (
            await db.execute(
                select(CountSession).where(CountSession.id == session_id)
            )
        ).scalar_one_or_none()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Conteo"

    # ── Styles ──
    header_fill = PatternFill("solid", fgColor="FF6B35")   # brand orange
    count_fill  = PatternFill("solid", fgColor="FFFDE7")   # soft yellow - column to fill
    system_fill = PatternFill("solid", fgColor="E8F5E9")   # soft green - system qty
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_font  = Font(bold=True, size=14)
    border_side = Side(style="thin", color="CCCCCC")
    thin_border = Border(
        left=border_side, right=border_side,
        top=border_side, bottom=border_side
    )

    # ── Title ──
    ws.merge_cells("A1:H1")
    ws["A1"] = f"CONTEO FÍSICO DE INVENTARIO — {session.name}"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = (
        f"Sesión ID: {session.id}  |  "
        f"Creada: {session.created_at.strftime('%d/%m/%Y %H:%M')}  |  "
        "INSTRUCCIONES: Llena solo la columna 'CONTEO FÍSICO' con la cantidad real contada"
    )
    ws["A2"].font = Font(italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 20

    # ── Headers row 3 ──
    headers = [
        ("A", "SKU", 18),
        ("B", "NOMBRE PRODUCTO", 40),
        ("C", "CATEGORÍA", 22),
        ("D", "COSTO UNIT.", 14),
        ("E", "STOCK SISTEMA", 14),
        ("F", "CONTEO FÍSICO ✏️", 16),
        ("G", "DIFERENCIA", 13),
        ("H", "NOTAS", 30),
    ]
    for col, label, width in headers:
        cell = ws[f"{col}3"]
        cell.value = label
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[col].width = width
    ws.row_dimensions[3].height = 30

    # ── Data rows ──
    # Get items sorted by category then name
    items = sorted(
        session.items,
        key=lambda i: (i.category_name or "ZZZ", i.product_name)
    )
    prev_cat = None
    data_start_row = 4

    for row_idx, item in enumerate(items, start=data_start_row):
        # Category separator row
        if item.category_name != prev_cat:
            prev_cat = item.category_name
            ws.merge_cells(f"A{row_idx}:H{row_idx}")
            cat_cell = ws[f"A{row_idx}"]
            cat_cell.value = f"  📦 {item.category_name or 'Sin categoría'}"
            cat_cell.fill = PatternFill("solid", fgColor="37474F")
            cat_cell.font = Font(bold=True, color="FFFFFF", size=10)
            cat_cell.alignment = Alignment(vertical="center")
            ws.row_dimensions[row_idx].height = 18
            row_idx += 1
            # Re-add item after separator
            # We'll write item on the next row
            ws.cell(row=row_idx, column=1).value = item.sku
            ws.cell(row=row_idx, column=2).value = item.product_name
            ws.cell(row=row_idx, column=3).value = item.category_name or ""
            ws.cell(row=row_idx, column=4).value = float(item.unit_cost)
            ws.cell(row=row_idx, column=5).value = item.system_qty
            ws.cell(row=row_idx, column=6).value = None  # To fill
            ws.cell(row=row_idx, column=7).value = None  # Formula
            ws.cell(row=row_idx, column=8).value = ""

            # Styles for data row
            for col in range(1, 9):
                c = ws.cell(row=row_idx, column=col)
                c.border = thin_border
                c.alignment = Alignment(vertical="center")
            ws.cell(row=row_idx, column=4).number_format = '#,##0.00'
            ws.cell(row=row_idx, column=5).fill = system_fill
            ws.cell(row=row_idx, column=5).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=6).fill = count_fill
            ws.cell(row=row_idx, column=6).font = Font(bold=True, size=12)
            ws.cell(row=row_idx, column=6).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=7).value = f"=IF(F{row_idx}=\"\",\"\",F{row_idx}-E{row_idx})"
            ws.cell(row=row_idx, column=7).number_format = '+0;-0;0'
            ws.row_dimensions[row_idx].height = 20
            continue

        # Normal data row (no category break)
        r = row_idx
        ws.cell(row=r, column=1).value = item.sku
        ws.cell(row=r, column=2).value = item.product_name
        ws.cell(row=r, column=3).value = item.category_name or ""
        ws.cell(row=r, column=4).value = float(item.unit_cost)
        ws.cell(row=r, column=5).value = item.system_qty
        ws.cell(row=r, column=6).value = None
        ws.cell(row=r, column=8).value = ""
        ws.cell(row=r, column=7).value = f"=IF(F{r}=\"\",\"\",F{r}-E{r})"
        ws.cell(row=r, column=7).number_format = '+0;-0;0'

        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.border = thin_border
            c.alignment = Alignment(vertical="center")
        ws.cell(row=r, column=4).number_format = '#,##0.00'
        ws.cell(row=r, column=5).fill = system_fill
        ws.cell(row=r, column=5).alignment = Alignment(horizontal="center")
        ws.cell(row=r, column=6).fill = count_fill
        ws.cell(row=r, column=6).font = Font(bold=True, size=12)
        ws.cell(row=r, column=6).alignment = Alignment(horizontal="center")
        ws.row_dimensions[r].height = 20

    # Freeze panes so header stays visible
    ws.freeze_panes = "A4"

    # ── Metadata sheet ──
    ws_meta = wb.create_sheet("Metadata")
    ws_meta["A1"] = "session_id"
    ws_meta["B1"] = str(session.id)
    ws_meta["A2"] = "session_name"
    ws_meta["B2"] = session.name
    ws_meta["A3"] = "created_at"
    ws_meta["B3"] = session.created_at.isoformat()

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in session.name)
    filename = f"conteo_{safe_name}_{session.created_at.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post(
    "/{session_id}/upload",
    response_model=UploadPreview,
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def upload_count_excel(
    session_id: uuid.UUID,
    db: DBSession,
    user: CurrentUser,
    file: UploadFile = File(...),
):
    """Sube el Excel diligenciado. Devuelve PREVIEW de diferencias sin aplicar nada todavía."""
    try:
        import openpyxl
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado")

    session = _session_or_404(
        (
            await db.execute(
                select(CountSession).where(CountSession.id == session_id)
            )
        ).scalar_one_or_none()
    )
    _require_status(session, "draft", "in_progress")

    content = await file.read()
    if not content:
        raise HTTPException(400, "Archivo vacío")
    if len(content) > 20 * 1024 * 1024:
        raise HTTPException(413, "Archivo > 20MB")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception as e:
        raise HTTPException(400, f"No se pudo leer el archivo Excel: {e}")

    ws = wb.active

    # Find header row: look for row with "SKU" in first column
    header_row = None
    sku_col = counted_col = system_col = notes_col = None
    for row in ws.iter_rows(max_row=10):
        for cell in row:
            val = str(cell.value or "").strip().upper()
            if val == "SKU":
                header_row = cell.row
                sku_col = cell.column
                break
        if header_row:
            break

    if header_row is None:
        raise HTTPException(400, "No se encontró columna 'SKU' en el archivo. Usa la plantilla oficial.")

    # Map header columns
    for cell in ws[header_row]:
        val = str(cell.value or "").strip().upper()
        if "CONTEO" in val or "FÍSICO" in val or "FISICO" in val:
            counted_col = cell.column
        elif "STOCK SISTEMA" in val or "SISTEMA" in val:
            system_col = cell.column
        elif "NOTA" in val:
            notes_col = cell.column

    if counted_col is None:
        raise HTTPException(
            400,
            "No se encontró la columna 'CONTEO FÍSICO'. "
            "Asegúrate de usar la plantilla descargada sin renombrar columnas.",
        )

    # Build SKU → CountItem map
    items_by_sku: dict[str, CountItem] = {it.sku.strip().upper(): it for it in session.items}

    rows_preview: list[UploadPreviewRow] = []
    counted_updates: list[tuple[CountItem, int, str | None]] = []
    not_found = 0

    for row in ws.iter_rows(min_row=header_row + 1):
        sku_cell = row[sku_col - 1]
        sku_val = str(sku_cell.value or "").strip()
        if not sku_val or sku_val.upper() == "SKU":
            continue

        counted_cell = row[counted_col - 1]
        raw_counted = counted_cell.value

        # Skip rows with no count entered
        if raw_counted is None or str(raw_counted).strip() == "":
            continue

        try:
            counted_qty = int(float(str(raw_counted)))
        except (ValueError, TypeError):
            continue

        if counted_qty < 0:
            continue

        notes_val = None
        if notes_col and notes_col <= len(row):
            notes_val = str(row[notes_col - 1].value or "").strip() or None

        item = items_by_sku.get(sku_val.upper())
        if item is None:
            not_found += 1
            rows_preview.append(UploadPreviewRow(
                sku=sku_val,
                product_name=f"[NO ENCONTRADO: {sku_val}]",
                category_name=None,
                system_qty=0,
                counted_qty=counted_qty,
                delta=0,
                value_impact=0,
                unit_cost=0,
                status="not_found",
            ))
            continue

        delta = counted_qty - item.system_qty
        value_impact = round(delta * float(item.unit_cost), 2)
        status = "ok" if delta == 0 else ("surplus" if delta > 0 else "shortage")

        rows_preview.append(UploadPreviewRow(
            sku=item.sku,
            product_name=item.product_name,
            category_name=item.category_name,
            system_qty=item.system_qty,
            counted_qty=counted_qty,
            delta=delta,
            value_impact=value_impact,
            unit_cost=float(item.unit_cost),
            status=status,
        ))
        counted_updates.append((item, counted_qty, notes_val))

    # Persist the counts into count_items (overwrite if re-uploaded)
    for item, counted_qty, notes in counted_updates:
        delta = counted_qty - item.system_qty
        item.counted_qty = counted_qty
        item.delta = delta
        item.value_impact = Decimal(str(round(delta * float(item.unit_cost), 2)))
        if notes:
            item.notes = notes

    # Update session status
    session.status = "in_progress"
    await db.commit()

    with_diff = sum(1 for r in rows_preview if r.status in ("surplus", "shortage"))
    total_value = sum(r.value_impact for r in rows_preview if r.status != "not_found")

    return UploadPreview(
        matched=len(counted_updates),
        not_found=not_found,
        with_difference=with_diff,
        total_value_impact=round(total_value, 2),
        rows=rows_preview,
    )


@router.post(
    "/{session_id}/apply",
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def apply_count_session(
    session_id: uuid.UUID,
    db: DBSession,
    user: CurrentUser,
):
    """Aplica los ajustes del conteo al stock real. IRREVERSIBLE.

    - Solo procesa items con counted_qty != None
    - Crea StockMovements tipo COUNT_ADJUST por cada diferencia
    - Actualiza inventory.stock directamente
    - Registra métricas en la sesión
    """
    session = _session_or_404(
        (
            await db.execute(
                select(CountSession).where(CountSession.id == session_id)
            )
        ).scalar_one_or_none()
    )
    _require_status(session, "in_progress")

    counted_items = [it for it in session.items if it.counted_qty is not None]
    if not counted_items:
        raise HTTPException(400, "No hay items contados. Sube el Excel primero.")

    loc = await _get_default_location(db)
    now = datetime.now(UTC)

    adjusted = 0
    pos_delta = 0
    neg_delta = 0
    value_impact_total = Decimal("0")

    for item in counted_items:
        delta = item.counted_qty - item.system_qty  # type: ignore[operator]
        if delta == 0:
            continue

        # Get or create stock row
        stock = (
            await db.execute(
                select(Stock)
                .where(Stock.product_id == item.product_id)
                .where(Stock.location_id == loc.id)
                .with_for_update()
            )
        ).scalar_one_or_none()

        if stock is None:
            if delta < 0:
                # Can't have negative stock where there's none
                item.delta = 0
                item.value_impact = Decimal("0")
                continue
            stock = Stock(
                product_id=item.product_id,
                location_id=loc.id,
                quantity=0,
                reserved=0,
            )
            db.add(stock)
            await db.flush()

        new_qty = max(0, stock.quantity + delta)
        actual_delta = new_qty - stock.quantity
        stock.quantity = new_qty

        movement = StockMovement(
            product_id=item.product_id,
            location_id=loc.id,
            movement_type="COUNT_ADJUST",
            quantity_delta=actual_delta,
            quantity_after=new_qty,
            unit_cost=float(item.unit_cost),
            reference_type="count_session",
            reference_id=session.id,
            notes=(
                f"Conteo físico '{session.name}' — "
                f"Sistema: {item.system_qty}, Contado: {item.counted_qty}"
            ),
            occurred_at=now,
            created_by=user.email,
        )
        db.add(movement)

        adjusted += 1
        if actual_delta > 0:
            pos_delta += actual_delta
        else:
            neg_delta += abs(actual_delta)
        impact = Decimal(str(round(actual_delta * float(item.unit_cost), 2)))
        value_impact_total += impact
        item.value_impact = impact
        item.delta = actual_delta

    # Update session metrics
    session.status = "applied"
    session.applied_at = now
    session.applied_by = user.email
    session.total_products_counted = len(counted_items)
    session.total_with_difference = adjusted
    session.total_positive_delta = pos_delta
    session.total_negative_delta = neg_delta
    session.total_value_impact = value_impact_total

    await db.commit()

    return {
        "status": "applied",
        "products_counted": len(counted_items),
        "products_adjusted": adjusted,
        "total_positive_delta": pos_delta,
        "total_negative_delta": neg_delta,
        "total_value_impact": float(value_impact_total),
    }


@router.delete(
    "/{session_id}",
    status_code=204,
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def delete_count_session(session_id: uuid.UUID, db: DBSession):
    """Elimina una sesión de conteo. Solo permitido si no ha sido aplicada."""
    session = _session_or_404(
        (
            await db.execute(
                select(CountSession).where(CountSession.id == session_id)
            )
        ).scalar_one_or_none()
    )
    _require_status(session, "draft", "in_progress")
    await db.delete(session)
    await db.commit()


@router.get(
    "/{session_id}/report",
    dependencies=[Depends(require_permission("inventory:adjust"))],
)
async def download_report(session_id: uuid.UUID, db: DBSession):
    """Descarga reporte Excel de diferencias de la sesión (solo sesiones aplicadas o in_progress)."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        raise HTTPException(500, "openpyxl no instalado")

    session = _session_or_404(
        (
            await db.execute(
                select(CountSession).where(CountSession.id == session_id)
            )
        ).scalar_one_or_none()
    )
    if session.status == "draft":
        raise HTTPException(400, "La sesión está en borrador. Sube el conteo primero.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte Diferencias"

    hdr_fill  = PatternFill("solid", fgColor="1A237E")
    hdr_font  = Font(bold=True, color="FFFFFF")
    plus_fill = PatternFill("solid", fgColor="E8F5E9")
    minus_fill = PatternFill("solid", fgColor="FFEBEE")
    ok_fill   = PatternFill("solid", fgColor="F5F5F5")

    headers = [
        "SKU", "Nombre", "Categoría", "Costo Unit.",
        "Stock Sistema", "Conteo Real", "Diferencia",
        "Impacto ($)", "Notas",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 15
    ws.column_dimensions["I"].width = 28

    items = sorted(session.items, key=lambda i: (i.category_name or "", i.product_name))
    for r, item in enumerate(items, start=2):
        if item.counted_qty is None:
            continue
        delta = item.delta or 0
        fill = plus_fill if delta > 0 else (minus_fill if delta < 0 else ok_fill)
        row_data = [
            item.sku,
            item.product_name,
            item.category_name or "",
            float(item.unit_cost),
            item.system_qty,
            item.counted_qty,
            delta,
            float(item.value_impact) if item.value_impact is not None else 0,
            item.notes or "",
        ]
        for col, val in enumerate(row_data, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill = fill
        ws.cell(row=r, column=4).number_format = "#,##0.00"
        ws.cell(row=r, column=8).number_format = "#,##0.00"

    # Summary row
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value="RESUMEN").font = Font(bold=True)
    ws.cell(row=last_row, column=5, value=f"=SUM(E2:E{ws.max_row - 1})")
    ws.cell(row=last_row, column=6, value=f"=SUM(F2:F{ws.max_row - 1})")
    ws.cell(row=last_row, column=7, value=f"=SUM(G2:G{ws.max_row - 1})")
    ws.cell(row=last_row, column=8, value=f"=SUM(H2:H{ws.max_row - 1})")
    ws.cell(row=last_row, column=8).number_format = "#,##0.00"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in " -_" else "_" for c in session.name)
    filename = f"reporte_conteo_{safe_name}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
