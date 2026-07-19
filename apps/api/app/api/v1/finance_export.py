"""Exportación Excel financiero ejecutivo — Bigotes y Paticas.

Endpoint: GET /v1/finance/export-excel?months=12
Genera un workbook de 5 hojas con análisis IA via Claude Haiku (OpenRouter).
"""

from __future__ import annotations

import io
import json
import os
from datetime import date, datetime, timedelta
from typing import Any
from zoneinfo import ZoneInfo

import httpx
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.deps import CurrentUser, DBSession, require_permission
from app.models.ops import LegacyIdMap
from sqlalchemy import select

export_router = APIRouter(prefix="/finance", tags=["finance"])

# ═══════════════════════════════════════════════════════
#  PALETA & ESTILOS
# ═══════════════════════════════════════════════════════

_TEAL_DARK  = "0D4A45"
_TEAL_MID   = "187F77"
_TEAL_LIGHT = "E0F2F1"
_GREEN_TXT  = "166534"
_GREEN_BG   = "DCFCE7"
_RED_TXT    = "991B1B"
_RED_BG     = "FEE2E2"
_ALT_ROW    = "F0FDF4"
_TOTAL_BG   = "D1FAE5"
_GRAY_BG    = "F8FAFC"
_WHITE      = "FFFFFF"
_BLACK      = "111827"
_PESO       = "#,##0"
_PESO_DEC   = "#,##0.00"
_PCT        = '0.0"%"'

_GASTOS_FIJOS = {"Arriendo", "Nómina", "Servicios", "Impuestos"}

_MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}


def _f(v: Any, default: float = 0.0) -> float:
    try:
        return float(v) if v is not None else default
    except (TypeError, ValueError):
        return default


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _border(color: str = "D1D5DB", style: str = "thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _font(bold: bool = False, size: int = 11, color: str = "111827", italic: bool = False) -> Font:
    return Font(bold=bold, size=size, color=color, italic=italic, name="Calibri")


def _w(ws, row: int, col: int, value, *, bold=False, size=11, fg=_BLACK, bg=None,
        align="left", wrap=False, fmt=None, border=True, italic=False):
    """Escribe una celda con estilos."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = _font(bold=bold, size=size, color=fg, italic=italic)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fmt:
        cell.number_format = fmt
    if border:
        cell.border = _border()
    return cell


def _col_w(ws, col: int, width: float) -> None:
    ws.column_dimensions[get_column_letter(col)].width = width


def _row_h(ws, row: int, height: float) -> None:
    ws.row_dimensions[row].height = height


def _merge(ws, r1, c1, r2, c2) -> None:
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _section_title(ws, row: int, col: int, ncols: int, title: str, subtitle: str = "") -> int:
    """Escribe un título de sección con merge y devuelve la siguiente fila."""
    _merge(ws, row, col, row, col + ncols - 1)
    cell = ws.cell(row=row, column=col, value=title)
    cell.font = _font(bold=True, size=13, color=_WHITE)
    cell.fill = _fill(_TEAL_MID)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = _border(_TEAL_DARK, "medium")
    _row_h(ws, row, 24)
    row += 1
    if subtitle:
        _merge(ws, row, col, row, col + ncols - 1)
        c = ws.cell(row=row, column=col, value=subtitle)
        c.font = _font(italic=True, size=10, color="4B5563")
        c.fill = _fill(_TEAL_LIGHT)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = _border()
        _row_h(ws, row, 16)
        row += 1
    return row


def _table_header(ws, row: int, cols: list[tuple[int, str]]) -> None:
    """Escribe una fila de cabecera de tabla."""
    for col, label in cols:
        cell = ws.cell(row=row, column=col, value=label)
        cell.font = _font(bold=True, size=10, color=_WHITE)
        cell.fill = _fill(_TEAL_DARK)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = _border(_TEAL_DARK)
    _row_h(ws, row, 20)


# ═══════════════════════════════════════════════════════
#  DATOS: FETCH FROM DB
# ═══════════════════════════════════════════════════════


def _months_ago_first_day(months: int) -> date:
    """Retorna el primer día del mes que está (months-1) meses antes del actual."""
    today = date.today()
    m = today.month - (months - 1)
    y = today.year
    while m <= 0:
        m += 12
        y -= 1
    return date(y, m, 1)


async def _fetch_monthly_sales(db: DBSession, months: int) -> list[dict]:
    """Retorna ventas mes a mes de los últimos N meses.

    IMPORTANTE: revenue y COGS se calculan en CTEs separadas para evitar que el
    LEFT JOIN con order_items multiplique grand_total por el número de ítems.
    La fecha usa zona horaria Colombia para que los meses coincidan con el admin.
    """
    since = _months_ago_first_day(months)
    rows = (
        await db.execute(
            text("""
            WITH revenue AS (
              SELECT
                DATE_TRUNC('month', o.occurred_at AT TIME ZONE 'America/Bogota') AS month,
                COALESCE(SUM(o.grand_total), 0)::float AS revenue,
                COUNT(o.id)::int                       AS orders
              FROM sales.orders o
              WHERE o.occurred_at >= :since
                AND COALESCE(o.status, '') NOT IN ('cancelled', 'refunded')
              GROUP BY DATE_TRUNC('month', o.occurred_at AT TIME ZONE 'America/Bogota')
            ),
            cogs_agg AS (
              SELECT
                DATE_TRUNC('month', o.occurred_at AT TIME ZONE 'America/Bogota') AS month,
                COALESCE(SUM(oi.unit_cost * oi.quantity), 0)::float AS cogs
              FROM sales.orders o
              JOIN sales.order_items oi ON oi.order_id = o.id
              WHERE o.occurred_at >= :since
                AND COALESCE(o.status, '') NOT IN ('cancelled', 'refunded')
              GROUP BY DATE_TRUNC('month', o.occurred_at AT TIME ZONE 'America/Bogota')
            )
            SELECT
              r.month,
              r.revenue,
              COALESCE(c.cogs, 0) AS cogs,
              r.orders
            FROM revenue r
            LEFT JOIN cogs_agg c ON c.month = r.month
            ORDER BY r.month
            """),
            {"since": since},
        )
    ).all()
    return [
        {
            "year_month": r.month.strftime("%Y-%m"),
            "month_label": f"{_MESES_ES[r.month.month]} {r.month.year}",
            "revenue": _f(r.revenue),
            "cogs": _f(r.cogs),
            "orders": int(r.orders or 0),
        }
        for r in rows
    ]


async def _fetch_all_expenses(db: DBSession, months: int) -> list[dict]:
    """Retorna todos los gastos de los últimos N meses."""
    since = _months_ago_first_day(months)
    rows = (await db.execute(select(LegacyIdMap).where(LegacyIdMap.entity == "gasto"))).scalars().all()
    result = []
    for r in rows:
        e = r.extra or {}
        raw_date = str(e.get("fecha", ""))[:10]
        try:
            d = datetime.fromisoformat(raw_date).date() if raw_date else None
        except Exception:
            d = None
        if d and d < since:
            continue
        monto = _f(e.get("monto"))
        cat = str(e.get("categoria", "Otros")) or "Otros"
        result.append({
            "fecha": raw_date or "",
            "year_month": raw_date[:7] if raw_date else "",
            "tipo": str(e.get("tipo", "")),
            "categoria": cat,
            "descripcion": str(e.get("descripcion", "")),
            "metodo_pago": str(e.get("metodo_pago", "")),
            "banco_origen": str(e.get("banco_origen", "")),
            "monto": monto,
            "tipo_gasto": "Fijo" if cat in _GASTOS_FIJOS else "Variable",
        })
    result.sort(key=lambda x: x["fecha"], reverse=True)
    return result


async def _fetch_inventory_value(db: DBSession) -> dict:
    """Retorna el valor del inventario actual."""
    row = (
        await db.execute(
            text("""
            SELECT
              COUNT(DISTINCT p.id)::int           AS total_products,
              COALESCE(SUM(s.quantity * p.cost), 0)::float AS value_cost,
              COALESCE(SUM(s.quantity * p.price), 0)::float AS value_price,
              SUM(CASE WHEN s.quantity = 0 THEN 1 ELSE 0 END)::int AS out_of_stock,
              SUM(CASE WHEN s.quantity > 0 AND s.quantity <= s.reorder_point THEN 1 ELSE 0 END)::int AS low_stock
            FROM catalog.products p
            LEFT JOIN inventory.stock s ON s.product_id = p.id
            WHERE p.deleted_at IS NULL AND p.is_active = true
            """)
        )
    ).one()
    return {
        "total_products": int(row.total_products or 0),
        "value_cost": _f(row.value_cost),
        "value_price": _f(row.value_price),
        "out_of_stock": int(row.out_of_stock or 0),
        "low_stock": int(row.low_stock or 0),
    }


async def _fetch_purchases_monthly(db: DBSession, months: int) -> dict[str, float]:
    """Retorna compras a proveedores mes a mes (monto total por mes)."""
    since = _months_ago_first_day(months)
    rows = (
        await db.execute(
            text("""
            SELECT
              DATE_TRUNC('month', purchased_at AT TIME ZONE 'America/Bogota') AS month,
              COALESCE(SUM(total), 0)::float AS total
            FROM purchasing.purchases
            WHERE purchased_at >= :since
              AND status IN ('confirmed', 'received')
            GROUP BY DATE_TRUNC('month', purchased_at AT TIME ZONE 'America/Bogota')
            ORDER BY month
            """),
            {"since": since},
        )
    ).all()
    return {r.month.strftime("%Y-%m"): _f(r.total) for r in rows}


async def _fetch_purchases_detail(db: DBSession, months: int) -> list[dict]:
    """Retorna detalle de cada orden de compra a proveedores."""
    since = _months_ago_first_day(months)
    rows = (
        await db.execute(
            text("""
            SELECT
              folio,
              COALESCE(supplier_name, 'Sin proveedor') AS supplier_name,
              status,
              COALESCE(subtotal, 0)::float   AS subtotal,
              COALESCE(tax_amount, 0)::float AS tax_amount,
              COALESCE(total, 0)::float      AS total,
              COALESCE(payment_method, '')   AS payment_method,
              purchased_at AT TIME ZONE 'America/Bogota' AS purchased_at_col
            FROM purchasing.purchases
            WHERE purchased_at >= :since
              AND status IN ('confirmed', 'received')
            ORDER BY purchased_at DESC
            """),
            {"since": since},
        )
    ).all()
    result = []
    for r in rows:
        pa = r.purchased_at_col
        result.append({
            "folio": str(r.folio or ""),
            "supplier_name": str(r.supplier_name),
            "status": str(r.status or ""),
            "subtotal": _f(r.subtotal),
            "tax_amount": _f(r.tax_amount),
            "total": _f(r.total),
            "payment_method": str(r.payment_method or ""),
            "fecha": pa.strftime("%Y-%m-%d") if pa else "",
            "year_month": pa.strftime("%Y-%m") if pa else "",
            "month_label": f"{_MESES_ES[pa.month]} {pa.year}" if pa else "",
        })
    return result


async def _fetch_top_suppliers(db: DBSession, months: int) -> list[dict]:
    """Retorna los proveedores con mayor gasto en el período."""
    since = _months_ago_first_day(months)
    rows = (
        await db.execute(
            text("""
            SELECT
              COALESCE(supplier_name, 'Sin proveedor') AS supplier_name,
              COUNT(*)::int                            AS num_compras,
              COALESCE(SUM(total), 0)::float           AS total_spend
            FROM purchasing.purchases
            WHERE purchased_at >= :since
              AND status IN ('confirmed', 'received')
            GROUP BY supplier_name
            ORDER BY total_spend DESC
            LIMIT 20
            """),
            {"since": since},
        )
    ).all()
    return [
        {"supplier_name": r.supplier_name, "num_compras": int(r.num_compras), "total_spend": _f(r.total_spend)}
        for r in rows
    ]


async def _fetch_revenue_by_method(db: DBSession, months: int) -> list[dict]:
    since = _months_ago_first_day(months)
    rows = (
        await db.execute(
            text("""
            SELECT p.method, COALESCE(SUM(p.amount), 0)::float AS total
            FROM sales.payments p
            JOIN sales.orders o ON o.id = p.order_id
            WHERE o.occurred_at >= :since
              AND COALESCE(o.status,'') NOT IN ('cancelled','refunded')
            GROUP BY p.method
            ORDER BY total DESC
            """),
            {"since": since},
        )
    ).all()
    return [{"method": r.method or "Sin método", "total": _f(r.total)} for r in rows]


# ═══════════════════════════════════════════════════════
#  IA: ANÁLISIS EJECUTIVO VIA OPENROUTER
# ═══════════════════════════════════════════════════════


async def _get_ai_analysis(summary: dict) -> dict:
    key = os.environ.get("OPENROUTER_API_KEY", "")
    fallback = {
        "diagnostico": "Análisis IA no disponible (clave OPENROUTER_API_KEY no configurada).",
        "fortalezas": ["Datos de ventas registrados correctamente.", "Sistema de inventario activo."],
        "riesgos": ["Verifique la configuración de la API de IA."],
        "recomendaciones": ["Configure OPENROUTER_API_KEY para habilitar el análisis IA."],
        "proyeccion": "No disponible.",
    }
    if not key:
        return fallback

    prompt = f"""Eres un analista financiero experto en retail de mascotas en Colombia.
Analiza los siguientes datos financieros REALES de Bigotes y Paticas (pet shop ubicado en Mall Zamara Plaza, Dosquebradas, Risaralda) y entrega un análisis ejecutivo tipo informe para inversionistas.

CONTEXTO DEL NEGOCIO:
- Tienda física + tienda online (bigotesypaticas.com)
- Portal de fidelización (mi.bigotesypaticas.com)
- Venta de alimentos, accesorios, medicamentos y snacks para mascotas
- Mercado: Pereira y Dosquebradas

DATOS FINANCIEROS (valores en COP colombiano):
{json.dumps(summary, ensure_ascii=False, indent=2)}

Responde en JSON estricto sin markdown ni texto adicional:
{{
  "diagnostico": "Párrafo de 4-5 oraciones describiendo la salud financiera del negocio, tendencias clave y posición competitiva actual",
  "fortalezas": ["fortaleza financiera 1 con dato concreto", "fortaleza 2 con dato concreto", "fortaleza 3 con dato concreto"],
  "riesgos": ["riesgo o área de mejora 1 con dato", "riesgo 2 con dato", "riesgo 3 con dato"],
  "recomendaciones": ["Acción específica 1: descripción + impacto esperado en COP o %", "Acción 2: descripción + impacto", "Acción 3: descripción + impacto"],
  "proyeccion": "Proyección fundamentada para los próximos 3 meses basada en la tendencia histórica, con cifras estimadas de revenue y rentabilidad"
}}"""

    try:
        async with httpx.AsyncClient(timeout=60) as client:
            resp = await client.post(
                "https://openrouter.ai/api/v1/chat/completions",
                headers={
                    "Authorization": f"Bearer {key}",
                    "HTTP-Referer": "https://bigotesypaticas.com",
                    "X-Title": "B&P Finance Report",
                },
                json={
                    "model": "anthropic/claude-haiku-4-5",
                    "messages": [{"role": "user", "content": prompt}],
                    "max_tokens": 2500,
                    "temperature": 0.2,
                },
            )
            resp.raise_for_status()
            raw = resp.json()["choices"][0]["message"]["content"].strip()
            if raw.startswith("```"):
                raw = raw.split("```", 2)[1]
                if raw.startswith("json"):
                    raw = raw[4:]
                if "```" in raw:
                    raw = raw.rsplit("```", 1)[0]
            return json.loads(raw.strip())
    except Exception:
        return fallback


# ═══════════════════════════════════════════════════════
#  EXCEL: CONSTRUCCIÓN DE HOJAS
# ═══════════════════════════════════════════════════════


def _sheet_summary(
    ws,
    months_data: list[dict],
    monthly_exp: dict[str, float],
    inv: dict,
    by_method: list[dict],
    ai: dict,
    months: int,
    generated_at: str,
) -> None:
    """Hoja 1: Resumen Ejecutivo."""

    # ── Columnas ─────────────────────────────────────
    _col_w(ws, 1, 3)
    _col_w(ws, 2, 28)
    _col_w(ws, 3, 22)
    _col_w(ws, 4, 22)
    _col_w(ws, 5, 22)
    _col_w(ws, 6, 22)
    _col_w(ws, 7, 3)

    # ── Encabezado corporativo ────────────────────────
    _merge(ws, 1, 2, 1, 6)
    ws.row_dimensions[1].height = 10

    _merge(ws, 2, 2, 2, 6)
    c = ws.cell(row=2, column=2, value="BIGOTES Y PATICAS")
    c.font = Font(bold=True, size=22, color=_WHITE, name="Calibri")
    c.fill = _fill(_TEAL_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center")
    _row_h(ws, 2, 36)

    _merge(ws, 3, 2, 3, 6)
    c = ws.cell(row=3, column=2, value="Informe Financiero Ejecutivo — Uso Confidencial")
    c.font = Font(bold=False, size=12, color=_WHITE, italic=True, name="Calibri")
    c.fill = _fill(_TEAL_MID)
    c.alignment = Alignment(horizontal="left", vertical="center")
    _row_h(ws, 3, 22)

    _merge(ws, 4, 2, 4, 6)
    c = ws.cell(row=4, column=2, value=f"Período analizado: últimos {months} meses  |  Generado: {generated_at}")
    c.font = Font(size=9, color="4B5563", italic=True, name="Calibri")
    c.fill = _fill(_TEAL_LIGHT)
    c.alignment = Alignment(horizontal="left", vertical="center")
    _row_h(ws, 4, 16)

    ws.row_dimensions[5].height = 8

    # ── Calcular totales ──────────────────────────────
    total_revenue = sum(m["revenue"] for m in months_data)
    total_cogs = sum(m["cogs"] for m in months_data)
    total_gross = total_revenue - total_cogs
    total_expenses = sum(monthly_exp.values())
    total_net = total_gross - total_expenses
    total_orders = sum(m["orders"] for m in months_data)
    avg_ticket = total_revenue / total_orders if total_orders > 0 else 0
    gross_margin = total_gross / total_revenue * 100 if total_revenue > 0 else 0
    net_margin = total_net / total_revenue * 100 if total_revenue > 0 else 0

    # Mes actual vs mes anterior
    sorted_months = sorted(months_data, key=lambda x: x["year_month"])
    cur = sorted_months[-1] if sorted_months else {}
    prev = sorted_months[-2] if len(sorted_months) >= 2 else {}
    cur_rev = cur.get("revenue", 0)
    prev_rev = prev.get("revenue", 0)
    rev_delta = (cur_rev - prev_rev) / prev_rev * 100 if prev_rev else 0
    cur_exp = monthly_exp.get(cur.get("year_month", ""), 0)
    prev_exp = monthly_exp.get(prev.get("year_month", ""), 0)
    cur_net = (cur_rev - cur.get("cogs", 0)) - cur_exp
    prev_net = (prev_rev - prev.get("cogs", 0)) - prev_exp

    # ── KPIs principales ──────────────────────────────
    row = 6
    kpis = [
        ("Ingresos Totales", total_revenue, _PESO, None),
        ("Costo de Ventas (COGS)", total_cogs, _PESO, None),
        ("Utilidad Bruta", total_gross, _PESO, _GREEN_BG if total_gross >= 0 else _RED_BG),
        ("Margen Bruto", gross_margin, _PCT, None),
        ("Gastos Operativos", total_expenses, _PESO, None),
        ("Utilidad Neta", total_net, _PESO, _GREEN_BG if total_net >= 0 else _RED_BG),
        ("Margen Neto", net_margin, _PCT, None),
        ("Pedidos Totales", total_orders, "#,##0", None),
        ("Ticket Promedio", avg_ticket, _PESO_DEC, None),
        ("Valor Inventario (costo)", inv.get("value_cost", 0), _PESO, None),
    ]

    row = _section_title(ws, row, 2, 5, "📊  KPIs FINANCIEROS DEL PERÍODO")
    _table_header(ws, row, [(2, "Indicador"), (3, "Valor"), (4, "Mes Actual"), (5, "Mes Anterior"), (6, "Variación")])
    _row_h(ws, row, 20)
    row += 1

    cur_month_label = cur.get("month_label", "")
    prev_month_label = prev.get("month_label", "")

    kpi_rows = [
        ("Ingresos", total_revenue, cur_rev, prev_rev,
         (cur_rev - prev_rev) / prev_rev * 100 if prev_rev else 0, _PESO),
        ("Utilidad Bruta", total_gross, cur_rev - cur.get("cogs", 0),
         prev_rev - prev.get("cogs", 0),
         ((cur_rev - cur.get("cogs", 0)) - (prev_rev - prev.get("cogs", 0))) / (prev_rev - prev.get("cogs", 1)) * 100
         if (prev_rev - prev.get("cogs", 0)) != 0 else 0, _PESO),
        ("Gastos Operativos", total_expenses, cur_exp, prev_exp,
         (cur_exp - prev_exp) / prev_exp * 100 if prev_exp else 0, _PESO),
        ("Utilidad Neta", total_net, cur_net, prev_net,
         (cur_net - prev_net) / abs(prev_net) * 100 if prev_net != 0 else 0, _PESO),
        ("Margen Neto %", net_margin,
         (cur_net / cur_rev * 100) if cur_rev else 0,
         (prev_net / prev_rev * 100) if prev_rev else 0,
         0, _PCT),
        ("Pedidos", total_orders, cur.get("orders", 0), prev.get("orders", 0),
         (cur.get("orders", 0) - prev.get("orders", 0)) / prev.get("orders", 1) * 100 if prev.get("orders") else 0, "#,##0"),
    ]

    for i, (label, total, cur_val, prev_val, delta, fmt) in enumerate(kpi_rows):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        _w(ws, row, 2, label, bold=True, size=10, bg=alt)
        _w(ws, row, 3, total, fmt=fmt, align="right", size=10, bg=alt, bold=True,
           fg=_GREEN_TXT if total >= 0 else _RED_TXT)
        _w(ws, row, 4, cur_val, fmt=fmt, align="right", size=10, bg=alt)
        _w(ws, row, 5, prev_val, fmt=fmt, align="right", size=10, bg=alt)
        delta_txt = f"+{delta:.1f}%" if delta >= 0 else f"{delta:.1f}%"
        delta_bg = _GREEN_BG if delta >= 0 else _RED_BG
        delta_fg = _GREEN_TXT if delta >= 0 else _RED_TXT
        _w(ws, row, 6, delta_txt, align="center", size=10, bg=delta_bg, fg=delta_fg, bold=True)
        _row_h(ws, row, 18)
        row += 1

    # Labels de mes en cabecera (sobreescribir con valores reales)
    ws.cell(row=row - len(kpi_rows) - 1, column=4).value = f"Mes Actual\n({cur_month_label})"
    ws.cell(row=row - len(kpi_rows) - 1, column=5).value = f"Mes Anterior\n({prev_month_label})"
    ws.cell(row=row - len(kpi_rows) - 1, column=4).alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
    ws.cell(row=row - len(kpi_rows) - 1, column=5).alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")

    row += 1

    # ── Análisis IA ────────────────────────────────────
    row = _section_title(ws, row, 2, 5, "🤖  ANÁLISIS INTELIGENTE (Claude Haiku IA)")

    # Diagnóstico
    _merge(ws, row, 2, row, 6)
    c = ws.cell(row=row, column=2, value="DIAGNÓSTICO DEL NEGOCIO")
    c.font = Font(bold=True, size=10, color=_TEAL_DARK, name="Calibri")
    c.fill = _fill(_TEAL_LIGHT)
    c.border = _border()
    _row_h(ws, row, 16)
    row += 1

    _merge(ws, row, 2, row, 6)
    c = ws.cell(row=row, column=2, value=ai.get("diagnostico", ""))
    c.font = Font(size=10, name="Calibri")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = _border()
    _row_h(ws, row, 72)
    row += 1

    # Fortalezas y Riesgos
    _merge(ws, row, 2, row, 3)
    c = ws.cell(row=row, column=2, value="✅  FORTALEZAS")
    c.font = Font(bold=True, size=10, color=_GREEN_TXT, name="Calibri")
    c.fill = _fill(_GREEN_BG)
    c.border = _border()

    _merge(ws, row, 4, row, 6)
    c = ws.cell(row=row, column=4, value="⚠️  ÁREAS DE MEJORA / RIESGOS")
    c.font = Font(bold=True, size=10, color=_RED_TXT, name="Calibri")
    c.fill = _fill(_RED_BG)
    c.border = _border()
    _row_h(ws, row, 16)
    row += 1

    fortalezas = ai.get("fortalezas", [])
    riesgos = ai.get("riesgos", [])
    max_items = max(len(fortalezas), len(riesgos), 1)
    for i in range(max_items):
        f_txt = f"• {fortalezas[i]}" if i < len(fortalezas) else ""
        r_txt = f"• {riesgos[i]}" if i < len(riesgos) else ""
        _merge(ws, row, 2, row, 3)
        c = ws.cell(row=row, column=2, value=f_txt)
        c.font = Font(size=10, name="Calibri")
        c.fill = _fill(_TEAL_LIGHT)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = _border()
        _merge(ws, row, 4, row, 6)
        c = ws.cell(row=row, column=4, value=r_txt)
        c.font = Font(size=10, name="Calibri")
        c.fill = _fill("FEF3C7")
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = _border()
        _row_h(ws, row, 40)
        row += 1

    row += 1

    # Recomendaciones
    _merge(ws, row, 2, row, 6)
    c = ws.cell(row=row, column=2, value="🎯  RECOMENDACIONES ESTRATÉGICAS")
    c.font = Font(bold=True, size=10, color=_WHITE, name="Calibri")
    c.fill = _fill(_TEAL_MID)
    c.border = _border()
    _row_h(ws, row, 16)
    row += 1

    for i, rec in enumerate(ai.get("recomendaciones", [])):
        _merge(ws, row, 2, row, 6)
        c = ws.cell(row=row, column=2, value=f"{i+1}. {rec}")
        c.font = Font(size=10, name="Calibri")
        c.fill = _fill(_ALT_ROW if i % 2 == 0 else _WHITE)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = _border()
        _row_h(ws, row, 36)
        row += 1

    row += 1

    # Proyección
    _merge(ws, row, 2, row, 6)
    c = ws.cell(row=row, column=2, value="🔮  PROYECCIÓN PRÓXIMOS 3 MESES")
    c.font = Font(bold=True, size=10, color=_WHITE, name="Calibri")
    c.fill = _fill(_TEAL_DARK)
    c.border = _border()
    _row_h(ws, row, 16)
    row += 1

    _merge(ws, row, 2, row, 6)
    c = ws.cell(row=row, column=2, value=ai.get("proyeccion", ""))
    c.font = Font(size=10, italic=True, name="Calibri")
    c.fill = _fill(_GRAY_BG)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = _border()
    _row_h(ws, row, 60)
    row += 1

    # Pie de página
    row += 1
    _merge(ws, row, 2, row, 6)
    c = ws.cell(row=row, column=2,
                value="Documento confidencial — Bigotes y Paticas | bigotesypaticas.com | Dosquebradas, Risaralda")
    c.font = Font(size=8, color="9CA3AF", italic=True, name="Calibri")
    c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "B6"
    ws.sheet_view.showGridLines = False


def _sheet_pl_monthly(
    ws,
    months_data: list[dict],
    monthly_exp: dict[str, float],
    purchases_monthly: dict[str, float],
) -> None:
    """Hoja 2: P&L Mensual — incluye Gastos Op. y Compras Proveedores separados."""

    _col_w(ws, 1, 3)
    _col_w(ws, 2, 22)
    _col_w(ws, 3, 18)
    _col_w(ws, 4, 18)
    _col_w(ws, 5, 18)
    _col_w(ws, 6, 12)
    _col_w(ws, 7, 18)
    _col_w(ws, 8, 18)   # Compras Proveedores
    _col_w(ws, 9, 18)
    _col_w(ws, 10, 12)
    _col_w(ws, 11, 13)
    _col_w(ws, 12, 13)

    ws.row_dimensions[1].height = 10
    row = 2
    row = _section_title(
        ws, row, 2, 10,
        "📈  P&L MENSUAL — ESTADO DE RESULTADOS",
        "Gastos Op. = gastos operativos (arriendo, nómina, etc.)  |  "
        "Compras Prov. = inversión en inventario (≠ COGS que es lo que ya se vendió)",
    )

    headers = [
        (2, "Mes"),
        (3, "Ingresos"),
        (4, "Costo Ventas\n(COGS)"),
        (5, "Utilidad Bruta"),
        (6, "Margen %"),
        (7, "Gastos Op."),
        (8, "Compras\nProveedores"),
        (9, "Utilidad Neta\n(sin compras)"),
        (10, "Margen\nNeto %"),
        (11, "Pedidos"),
        (12, "Ticket Prom."),
    ]
    _table_header(ws, row, headers)
    _row_h(ws, row, 30)
    row += 1

    totals = {k: 0.0 for k in ["revenue", "cogs", "gross", "expenses", "purchases", "net", "orders"]}
    sorted_months = sorted(months_data, key=lambda x: x["year_month"])

    for i, m in enumerate(sorted_months):
        ym = m["year_month"]
        revenue = m["revenue"]
        cogs = m["cogs"]
        gross = revenue - cogs
        expenses = monthly_exp.get(ym, 0.0)
        purchases = purchases_monthly.get(ym, 0.0)
        net = gross - expenses          # utilidad neta operativa (no incluye compras de inventario)
        orders = m["orders"]
        ticket = revenue / orders if orders > 0 else 0
        gross_pct = gross / revenue * 100 if revenue > 0 else 0
        net_pct = net / revenue * 100 if revenue > 0 else 0

        totals["revenue"] += revenue
        totals["cogs"] += cogs
        totals["gross"] += gross
        totals["expenses"] += expenses
        totals["purchases"] += purchases
        totals["net"] += net
        totals["orders"] += orders

        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        net_bg = _GREEN_BG if net >= 0 else _RED_BG
        net_fg = _GREEN_TXT if net >= 0 else _RED_TXT
        buy_bg = "FEF3C7" if purchases > 0 else alt  # amarillo si hay compras

        _w(ws, row, 2, m.get("month_label", ym), bold=True, size=10, bg=alt)
        _w(ws, row, 3, revenue, fmt=_PESO, align="right", size=10, bg=alt)
        _w(ws, row, 4, cogs, fmt=_PESO, align="right", size=10, bg=alt)
        _w(ws, row, 5, gross, fmt=_PESO, align="right", size=10, bg=alt,
           fg=_GREEN_TXT if gross >= 0 else _RED_TXT)
        _w(ws, row, 6, gross_pct, fmt=_PCT, align="center", size=10, bg=alt)
        _w(ws, row, 7, expenses, fmt=_PESO, align="right", size=10, bg=alt)
        _w(ws, row, 8, purchases, fmt=_PESO, align="right", size=10, bg=buy_bg,
           fg="92400E" if purchases > 0 else _BLACK, bold=purchases > 0)
        _w(ws, row, 9, net, fmt=_PESO, align="right", size=10, bg=net_bg, fg=net_fg, bold=True)
        _w(ws, row, 10, net_pct, fmt=_PCT, align="center", size=10, bg=net_bg, fg=net_fg)
        _w(ws, row, 11, orders, fmt="#,##0", align="center", size=10, bg=alt)
        _w(ws, row, 12, ticket, fmt=_PESO, align="right", size=10, bg=alt)
        _row_h(ws, row, 18)
        row += 1

    # Totales
    total_rev = totals["revenue"]
    total_gross = totals["gross"]
    total_net = totals["net"]
    total_orders = int(totals["orders"])
    total_ticket = total_rev / total_orders if total_orders > 0 else 0

    _w(ws, row, 2, "TOTAL / PROMEDIO", bold=True, size=11, bg=_TOTAL_BG, fg=_TEAL_DARK)
    _w(ws, row, 3, total_rev, fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 4, totals["cogs"], fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 5, total_gross, fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG,
       fg=_GREEN_TXT if total_gross >= 0 else _RED_TXT)
    avg_gm = total_gross / total_rev * 100 if total_rev > 0 else 0
    _w(ws, row, 6, avg_gm, fmt=_PCT, align="center", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 7, totals["expenses"], fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 8, totals["purchases"], fmt=_PESO, align="right", bold=True, size=11, bg="FEF3C7",
       fg="92400E")
    net_bg = _GREEN_BG if total_net >= 0 else _RED_BG
    _w(ws, row, 9, total_net, fmt=_PESO, align="right", bold=True, size=11, bg=net_bg,
       fg=_GREEN_TXT if total_net >= 0 else _RED_TXT)
    avg_nm = total_net / total_rev * 100 if total_rev > 0 else 0
    _w(ws, row, 10, avg_nm, fmt=_PCT, align="center", bold=True, size=11, bg=net_bg)
    _w(ws, row, 11, total_orders, fmt="#,##0", align="center", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 12, total_ticket, fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG)
    _row_h(ws, row, 22)

    ws.freeze_panes = "C4"
    ws.auto_filter.ref = f"B{row - len(sorted_months)}:L{row - 1}"
    ws.sheet_view.showGridLines = False


def _sheet_expenses(ws, all_expenses: list[dict], monthly_exp: dict[str, float]) -> None:
    """Hoja 3: Gastos Detallado."""

    _col_w(ws, 1, 3)
    _col_w(ws, 2, 14)
    _col_w(ws, 3, 16)
    _col_w(ws, 4, 18)
    _col_w(ws, 5, 35)
    _col_w(ws, 6, 16)
    _col_w(ws, 7, 16)
    _col_w(ws, 8, 12)

    ws.row_dimensions[1].height = 10
    row = 2

    # ── Resumen por categoría ─────────────────────────
    row = _section_title(ws, row, 2, 6, "💸  GASTOS OPERATIVOS — RESUMEN POR CATEGORÍA")
    _table_header(ws, row, [
        (2, "Categoría"), (3, "Tipo"), (4, "Total (COP)"), (5, "% del Total"), (6, "Registros"),
    ])
    _row_h(ws, row, 20)
    row += 1

    by_cat: dict[str, dict] = {}
    for e in all_expenses:
        cat = e["categoria"]
        if cat not in by_cat:
            by_cat[cat] = {"total": 0, "count": 0, "tipo": e["tipo_gasto"]}
        by_cat[cat]["total"] += e["monto"]
        by_cat[cat]["count"] += 1

    total_gastos = sum(v["total"] for v in by_cat.values())
    fijo_total = sum(v["total"] for v in by_cat.values() if v["tipo"] == "Fijo")
    var_total = sum(v["total"] for v in by_cat.values() if v["tipo"] == "Variable")

    sorted_cats = sorted(by_cat.items(), key=lambda x: -x[1]["total"])
    for i, (cat, info) in enumerate(sorted_cats):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        pct = info["total"] / total_gastos * 100 if total_gastos > 0 else 0
        _w(ws, row, 2, cat, bold=True, size=10, bg=alt)
        tipo_bg = _TEAL_LIGHT if info["tipo"] == "Fijo" else "FEF3C7"
        _w(ws, row, 3, info["tipo"], align="center", size=10, bg=tipo_bg,
           fg=_TEAL_DARK if info["tipo"] == "Fijo" else "92400E")
        _w(ws, row, 4, info["total"], fmt=_PESO, align="right", size=10, bg=alt, bold=True)
        _w(ws, row, 5, pct, fmt=_PCT, align="center", size=10, bg=alt)
        _w(ws, row, 6, info["count"], fmt="#,##0", align="center", size=10, bg=alt)
        _row_h(ws, row, 18)
        row += 1

    # Total
    _w(ws, row, 2, "TOTAL GASTOS", bold=True, size=11, bg=_TOTAL_BG, fg=_TEAL_DARK)
    _w(ws, row, 3, "", bg=_TOTAL_BG)
    _w(ws, row, 4, total_gastos, fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG, fg=_RED_TXT)
    _w(ws, row, 5, 100.0, fmt=_PCT, align="center", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 6, len(all_expenses), fmt="#,##0", align="center", bold=True, size=11, bg=_TOTAL_BG)
    _row_h(ws, row, 22)
    row += 2

    # ── Fijo vs Variable ─────────────────────────────
    row = _section_title(ws, row, 2, 5, "📊  GASTOS FIJOS vs VARIABLES")
    _table_header(ws, row, [(2, "Tipo"), (3, "Monto (COP)"), (4, "% Total"), (5, "Categorías incluidas")])
    _row_h(ws, row, 20)
    row += 1

    fijo_cats = ", ".join(sorted(_GASTOS_FIJOS))
    var_cats = ", ".join(sorted(set(c for c, v in by_cat.items() if v["tipo"] == "Variable")))

    _w(ws, row, 2, "Gastos Fijos", bold=True, size=10, bg=_TEAL_LIGHT, fg=_TEAL_DARK)
    _w(ws, row, 3, fijo_total, fmt=_PESO, align="right", bold=True, size=10, bg=_TEAL_LIGHT)
    fijo_pct = fijo_total / total_gastos * 100 if total_gastos else 0
    _w(ws, row, 4, fijo_pct, fmt=_PCT, align="center", size=10, bg=_TEAL_LIGHT)
    _w(ws, row, 5, fijo_cats, size=9, bg=_TEAL_LIGHT, wrap=True)
    _row_h(ws, row, 26)
    row += 1

    _w(ws, row, 2, "Gastos Variables", bold=True, size=10, bg="FEF3C7", fg="92400E")
    _w(ws, row, 3, var_total, fmt=_PESO, align="right", bold=True, size=10, bg="FEF3C7")
    var_pct = var_total / total_gastos * 100 if total_gastos else 0
    _w(ws, row, 4, var_pct, fmt=_PCT, align="center", size=10, bg="FEF3C7")
    _w(ws, row, 5, var_cats, size=9, bg="FEF3C7", wrap=True)
    _row_h(ws, row, 26)
    row += 2

    # ── Lista completa ────────────────────────────────
    row = _section_title(ws, row, 2, 7, "📋  DETALLE COMPLETO DE GASTOS")
    _table_header(ws, row, [
        (2, "Fecha"), (3, "Tipo"), (4, "Categoría"), (5, "Descripción"),
        (6, "Método Pago"), (7, "Fijo/Variable"), (8, "Monto (COP)"),
    ])
    _row_h(ws, row, 20)
    detail_start = row
    row += 1

    for i, e in enumerate(all_expenses):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        tipo_bg = _TEAL_LIGHT if e["tipo_gasto"] == "Fijo" else "FEF3C7"
        tipo_fg = _TEAL_DARK if e["tipo_gasto"] == "Fijo" else "92400E"
        _w(ws, row, 2, e["fecha"], size=9, bg=alt)
        _w(ws, row, 3, e["tipo"], size=9, bg=alt)
        _w(ws, row, 4, e["categoria"], bold=True, size=9, bg=alt)
        _w(ws, row, 5, e["descripcion"], size=9, bg=alt, wrap=True)
        _w(ws, row, 6, e["metodo_pago"], size=9, bg=alt, align="center")
        _w(ws, row, 7, e["tipo_gasto"], size=9, bg=tipo_bg, fg=tipo_fg, align="center", bold=True)
        _w(ws, row, 8, e["monto"], fmt=_PESO, align="right", size=9, bg=alt,
           fg=_RED_TXT, bold=True)
        _row_h(ws, row, 16)
        row += 1

    if all_expenses:
        ws.auto_filter.ref = f"B{detail_start + 1}:H{row - 1}"

    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False


def _sheet_income(ws, months_data: list[dict], by_method: list[dict]) -> None:
    """Hoja 4: Ingresos Análisis."""

    _col_w(ws, 1, 3)
    _col_w(ws, 2, 22)
    _col_w(ws, 3, 18)
    _col_w(ws, 4, 16)
    _col_w(ws, 5, 12)
    _col_w(ws, 6, 18)
    _col_w(ws, 7, 18)

    ws.row_dimensions[1].height = 10
    row = 2

    # ── Por método de pago ────────────────────────────
    row = _section_title(ws, row, 2, 5, "💳  INGRESOS POR MÉTODO DE PAGO")
    _table_header(ws, row, [
        (2, "Método de Pago"), (3, "Ingresos (COP)"), (4, "% del Total"), (5, "Ranking"),
    ])
    _row_h(ws, row, 20)
    row += 1

    total_by_method = sum(m["total"] for m in by_method)
    for i, m in enumerate(by_method):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        pct = m["total"] / total_by_method * 100 if total_by_method else 0
        _w(ws, row, 2, m["method"], bold=True, size=10, bg=alt)
        _w(ws, row, 3, m["total"], fmt=_PESO, align="right", size=10, bg=alt, bold=True)
        _w(ws, row, 4, pct, fmt=_PCT, align="center", size=10, bg=alt)
        _w(ws, row, 5, f"#{i+1}", align="center", size=10, bg=alt, bold=True, fg=_TEAL_DARK)
        _row_h(ws, row, 18)
        row += 1

    _w(ws, row, 2, "TOTAL INGRESOS", bold=True, size=11, bg=_TOTAL_BG, fg=_TEAL_DARK)
    _w(ws, row, 3, total_by_method, fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 4, 100.0, fmt=_PCT, align="center", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 5, "", bg=_TOTAL_BG)
    _row_h(ws, row, 22)
    row += 2

    # ── Tendencia mensual ─────────────────────────────
    row = _section_title(ws, row, 2, 5, "📅  TENDENCIA DE INGRESOS MENSUALES")
    _table_header(ws, row, [
        (2, "Mes"), (3, "Ingresos (COP)"), (4, "Pedidos"), (5, "Ticket Prom."),
        (6, "Var. vs Anterior"), (7, "Tendencia"),
    ])
    _col_w(ws, 7, 14)
    _row_h(ws, row, 20)
    row += 1

    sorted_months = sorted(months_data, key=lambda x: x["year_month"])
    for i, m in enumerate(sorted_months):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        orders = m["orders"]
        ticket = m["revenue"] / orders if orders > 0 else 0
        prev_rev = sorted_months[i - 1]["revenue"] if i > 0 else 0
        var_pct = (m["revenue"] - prev_rev) / prev_rev * 100 if prev_rev > 0 and i > 0 else 0
        tendencia = "↑ SUBE" if var_pct > 2 else ("↓ BAJA" if var_pct < -2 else "→ ESTABLE")
        tend_bg = _GREEN_BG if var_pct > 2 else (_RED_BG if var_pct < -2 else "FEF3C7")
        tend_fg = _GREEN_TXT if var_pct > 2 else (_RED_TXT if var_pct < -2 else "92400E")

        _w(ws, row, 2, m.get("month_label", m["year_month"]), bold=True, size=10, bg=alt)
        _w(ws, row, 3, m["revenue"], fmt=_PESO, align="right", size=10, bg=alt, bold=True)
        _w(ws, row, 4, orders, fmt="#,##0", align="center", size=10, bg=alt)
        _w(ws, row, 5, ticket, fmt=_PESO, align="right", size=10, bg=alt)
        var_txt = f"+{var_pct:.1f}%" if var_pct >= 0 else f"{var_pct:.1f}%" if i > 0 else "—"
        _w(ws, row, 6, var_txt, align="center", size=10,
           bg=tend_bg if i > 0 else alt, fg=tend_fg if i > 0 else _BLACK)
        _w(ws, row, 7, tendencia if i > 0 else "—", align="center", size=10,
           bg=tend_bg if i > 0 else alt, fg=tend_fg if i > 0 else _BLACK, bold=True)
        _row_h(ws, row, 18)
        row += 1

    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False


def _sheet_inventory(ws, inv: dict, generated_at: str) -> None:
    """Hoja 5: Inventario — snapshot actual."""

    _col_w(ws, 1, 3)
    _col_w(ws, 2, 32)
    _col_w(ws, 3, 22)
    _col_w(ws, 4, 22)

    ws.row_dimensions[1].height = 10
    row = 2

    row = _section_title(ws, row, 2, 3, "📦  INVENTARIO — SNAPSHOT ACTUAL",
                         f"Datos al {generated_at} (valor en libros, no incluye ajustes de mercado)")

    kpis = [
        ("Valor del Inventario a Costo", inv.get("value_cost", 0), _PESO, _TEAL_LIGHT, _TEAL_DARK),
        ("Valor del Inventario a Precio Venta", inv.get("value_price", 0), _PESO, _GREEN_BG, _GREEN_TXT),
        ("Margen Potencial Inventario", inv.get("value_price", 0) - inv.get("value_cost", 0), _PESO, "FEF3C7", "92400E"),
        ("Total Productos Activos", inv.get("total_products", 0), "#,##0", _GRAY_BG, _BLACK),
        ("Productos Sin Stock", inv.get("out_of_stock", 0), "#,##0", _RED_BG, _RED_TXT),
        ("Productos con Stock Bajo", inv.get("low_stock", 0), "#,##0", "FEF3C7", "92400E"),
    ]

    for label, value, fmt, bg, fg in kpis:
        _merge(ws, row, 2, row, 2)
        _w(ws, row, 2, label, bold=True, size=11, bg=bg, fg=fg)
        _merge(ws, row, 3, row, 4)
        c = ws.cell(row=row, column=3, value=value)
        c.font = Font(bold=True, size=14, color=fg, name="Calibri")
        c.fill = _fill(bg)
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.number_format = fmt
        c.border = _border()
        _row_h(ws, row, 28)
        row += 1

    row += 1
    _merge(ws, row, 2, row, 4)
    c = ws.cell(row=row, column=2,
                value="⚠️  Nota: El valor del inventario es un cálculo basado en stock actual × costo unitario registrado. "
                      "No incluye mermas, vencimientos o productos sin costo registrado.")
    c.font = Font(size=9, italic=True, color="6B7280", name="Calibri")
    c.fill = _fill("FFFBEB")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = _border()
    _row_h(ws, row, 42)

    ws.sheet_view.showGridLines = False


def _sheet_purchases(
    ws,
    purchases_detail: list[dict],
    purchases_monthly: dict[str, float],
    top_suppliers: list[dict],
    months_data: list[dict],
) -> None:
    """Hoja: Compras a Proveedores — separada de Gastos Operativos."""

    _col_w(ws, 1, 3)
    _col_w(ws, 2, 14)
    _col_w(ws, 3, 16)
    _col_w(ws, 4, 30)
    _col_w(ws, 5, 18)
    _col_w(ws, 6, 14)
    _col_w(ws, 7, 18)
    _col_w(ws, 8, 14)

    ws.row_dimensions[1].height = 10
    row = 2

    # ── Resumen mensual ───────────────────────────────
    row = _section_title(
        ws, row, 2, 6,
        "🛒  COMPRAS A PROVEEDORES — INVERSIÓN EN INVENTARIO",
        "Nota: Las compras aumentan el inventario (activo). El costo de lo vendido (COGS) "
        "aparece en el P&L cuando el producto se vende, no cuando se compra.",
    )

    _table_header(ws, row, [
        (2, "Mes"), (3, "Compras (COP)"), (4, "Ingresos mes"), (5, "% Compras/Ingr."), (6, "# Órdenes"),
    ])
    _row_h(ws, row, 20)
    row += 1

    # Índice de órdenes por mes
    orders_by_month: dict[str, int] = {}
    for p in purchases_detail:
        ym = p["year_month"]
        if ym:
            orders_by_month[ym] = orders_by_month.get(ym, 0) + 1

    revenue_by_month: dict[str, float] = {m["year_month"]: m["revenue"] for m in months_data}

    all_ym = sorted(set(list(purchases_monthly.keys()) + list(revenue_by_month.keys())))
    total_purchases = sum(purchases_monthly.values())
    total_revenue_pm = sum(revenue_by_month.values())

    for i, ym in enumerate(all_ym):
        purch = purchases_monthly.get(ym, 0.0)
        rev = revenue_by_month.get(ym, 0.0)
        ratio = purch / rev * 100 if rev > 0 else 0
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        n_orders = orders_by_month.get(ym, 0)

        try:
            y, m_ = int(ym[:4]), int(ym[5:7])
            label = f"{_MESES_ES[m_]} {y}"
        except Exception:
            label = ym

        _w(ws, row, 2, label, bold=True, size=10, bg=alt)
        _w(ws, row, 3, purch, fmt=_PESO, align="right", size=10, bg="FEF3C7" if purch > 0 else alt,
           fg="92400E" if purch > 0 else _BLACK, bold=purch > 0)
        _w(ws, row, 4, rev, fmt=_PESO, align="right", size=10, bg=alt)
        ratio_bg = _RED_BG if ratio > 80 else ("FEF3C7" if ratio > 50 else alt)
        _w(ws, row, 5, ratio, fmt=_PCT, align="center", size=10, bg=ratio_bg)
        _w(ws, row, 6, n_orders, fmt="#,##0", align="center", size=10, bg=alt)
        _row_h(ws, row, 18)
        row += 1

    # Fila totales
    total_ratio = total_purchases / total_revenue_pm * 100 if total_revenue_pm > 0 else 0
    _w(ws, row, 2, "TOTAL PERÍODO", bold=True, size=11, bg=_TOTAL_BG, fg=_TEAL_DARK)
    _w(ws, row, 3, total_purchases, fmt=_PESO, align="right", bold=True, size=11, bg="FEF3C7", fg="92400E")
    _w(ws, row, 4, total_revenue_pm, fmt=_PESO, align="right", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 5, total_ratio, fmt=_PCT, align="center", bold=True, size=11, bg=_TOTAL_BG)
    _w(ws, row, 6, len(purchases_detail), fmt="#,##0", align="center", bold=True, size=11, bg=_TOTAL_BG)
    _row_h(ws, row, 22)
    row += 2

    # ── Top proveedores ───────────────────────────────
    row = _section_title(ws, row, 2, 5, "🏭  TOP PROVEEDORES POR GASTO")
    _table_header(ws, row, [
        (2, "Proveedor"), (3, "Total Comprado (COP)"), (4, "# Órdenes"), (5, "% del Total"),
    ])
    _row_h(ws, row, 20)
    row += 1

    for i, s in enumerate(top_suppliers):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        pct = s["total_spend"] / total_purchases * 100 if total_purchases > 0 else 0
        _w(ws, row, 2, s["supplier_name"], bold=True, size=10, bg=alt)
        _w(ws, row, 3, s["total_spend"], fmt=_PESO, align="right", size=10, bg=alt, bold=True)
        _w(ws, row, 4, s["num_compras"], fmt="#,##0", align="center", size=10, bg=alt)
        _w(ws, row, 5, pct, fmt=_PCT, align="center", size=10, bg=alt)
        _row_h(ws, row, 18)
        row += 1

    row += 1

    # ── Detalle completo ──────────────────────────────
    row = _section_title(ws, row, 2, 7, "📋  DETALLE DE ÓRDENES DE COMPRA")
    _table_header(ws, row, [
        (2, "Fecha"), (3, "Folio"), (4, "Proveedor"),
        (5, "Subtotal"), (6, "IVA"), (7, "Total"), (8, "Método Pago"),
    ])
    _row_h(ws, row, 20)
    detail_start = row
    row += 1

    for i, p in enumerate(purchases_detail):
        alt = _ALT_ROW if i % 2 == 0 else _WHITE
        _w(ws, row, 2, p["fecha"], size=9, bg=alt)
        _w(ws, row, 3, p["folio"], size=9, bg=alt)
        _w(ws, row, 4, p["supplier_name"], bold=True, size=9, bg=alt)
        _w(ws, row, 5, p["subtotal"], fmt=_PESO, align="right", size=9, bg=alt)
        _w(ws, row, 6, p["tax_amount"], fmt=_PESO, align="right", size=9, bg=alt)
        _w(ws, row, 7, p["total"], fmt=_PESO, align="right", size=9, bg="FEF3C7",
           fg="92400E", bold=True)
        _w(ws, row, 8, p["payment_method"], align="center", size=9, bg=alt)
        _row_h(ws, row, 16)
        row += 1

    if purchases_detail:
        ws.auto_filter.ref = f"B{detail_start + 1}:H{row - 1}"

    ws.freeze_panes = "C4"
    ws.sheet_view.showGridLines = False


# ═══════════════════════════════════════════════════════
#  ENDPOINT PRINCIPAL
# ═══════════════════════════════════════════════════════


@export_router.get(
    "/export-excel",
    dependencies=[Depends(require_permission("finance:read"))],
    response_class=StreamingResponse,
    summary="Exporta informe financiero ejecutivo (.xlsx) con análisis IA",
)
async def export_finance_excel(
    db: DBSession,
    user: CurrentUser,
    months: int = Query(12, ge=1, le=24, description="Meses a analizar"),
):
    tz_col = ZoneInfo("America/Bogota")
    now_col = datetime.now(tz_col)
    generated_at = now_col.strftime("%d/%m/%Y %H:%M")

    # ── 1. Fetch data ─────────────────────────────────
    (
        monthly_sales,
        all_expenses,
        inv,
        by_method,
        purchases_monthly,
        purchases_detail,
        top_suppliers,
    ) = (
        await _fetch_monthly_sales(db, months),
        await _fetch_all_expenses(db, months),
        await _fetch_inventory_value(db),
        await _fetch_revenue_by_method(db, months),
        await _fetch_purchases_monthly(db, months),
        await _fetch_purchases_detail(db, months),
        await _fetch_top_suppliers(db, months),
    )

    # Group expenses by month
    monthly_exp: dict[str, float] = {}
    for e in all_expenses:
        ym = e["year_month"]
        if ym:
            monthly_exp[ym] = monthly_exp.get(ym, 0.0) + e["monto"]

    # ── 2. Build AI summary payload ───────────────────
    total_rev = sum(m["revenue"] for m in monthly_sales)
    total_cogs = sum(m["cogs"] for m in monthly_sales)
    total_exp = sum(monthly_exp.values())
    total_gross = total_rev - total_cogs
    total_net = total_gross - total_exp
    total_orders = sum(m["orders"] for m in monthly_sales)

    exp_by_cat: dict[str, float] = {}
    for e in all_expenses:
        exp_by_cat[e["categoria"]] = exp_by_cat.get(e["categoria"], 0.0) + e["monto"]

    sorted_months_ai = sorted(monthly_sales, key=lambda x: x["year_month"])
    trend_summary = [
        {
            "mes": m["month_label"],
            "ingresos": m["revenue"],
            "pedidos": m["orders"],
        }
        for m in sorted_months_ai
    ]

    ai_payload = {
        "periodo_analizado_meses": months,
        "ingresos_totales_COP": round(total_rev, 0),
        "costo_de_ventas_COP": round(total_cogs, 0),
        "utilidad_bruta_COP": round(total_gross, 0),
        "margen_bruto_pct": round(total_gross / total_rev * 100, 1) if total_rev else 0,
        "gastos_operativos_COP": round(total_exp, 0),
        "utilidad_neta_COP": round(total_net, 0),
        "margen_neto_pct": round(total_net / total_rev * 100, 1) if total_rev else 0,
        "total_pedidos": total_orders,
        "ticket_promedio_COP": round(total_rev / total_orders, 0) if total_orders else 0,
        "valor_inventario_costo_COP": round(inv.get("value_cost", 0), 0),
        "productos_sin_stock": inv.get("out_of_stock", 0),
        "gastos_por_categoria_COP": {k: round(v, 0) for k, v in sorted(exp_by_cat.items(), key=lambda x: -x[1])},
        "compras_proveedores_COP": round(sum(purchases_monthly.values()), 0),
        "tendencia_mensual": trend_summary,
        "ingresos_por_metodo": {m["method"]: round(m["total"], 0) for m in by_method},
    }

    # ── 3. IA Analysis ────────────────────────────────
    ai_analysis = await _get_ai_analysis(ai_payload)

    # ── 4. Build Excel ────────────────────────────────
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("Resumen Ejecutivo")
    _sheet_summary(ws1, monthly_sales, monthly_exp, inv, by_method, ai_analysis, months, generated_at)

    ws2 = wb.create_sheet("P&L Mensual")
    _sheet_pl_monthly(ws2, monthly_sales, monthly_exp, purchases_monthly)

    ws3 = wb.create_sheet("Gastos Operativos")
    _sheet_expenses(ws3, all_expenses, monthly_exp)

    ws4 = wb.create_sheet("Compras Proveedores")
    _sheet_purchases(ws4, purchases_detail, purchases_monthly, top_suppliers, monthly_sales)

    ws5 = wb.create_sheet("Ingresos Análisis")
    _sheet_income(ws5, monthly_sales, by_method)

    ws6 = wb.create_sheet("Inventario")
    _sheet_inventory(ws6, inv, generated_at)

    # ── 5. Serialize & return ─────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"BigotesyPaticas_Informe_Financiero_{now_col.strftime('%Y%m%d')}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Cache-Control": "no-store",
        },
    )
