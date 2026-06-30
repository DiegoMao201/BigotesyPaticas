"""Endpoints de catálogo: productos, marcas, categorías."""

from __future__ import annotations

import io
import math
import uuid
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from slugify import slugify
from sqlalchemy import func, or_, select, update

from app.deps import DBSession, require_permission
from app.models.catalog import Brand, Category, Product, ProductReview
from app.models.crm import Customer
from app.models.inventory import Stock
from app.models.purchasing import Supplier, SupplierSkuMap
from app.schemas.catalog import (
    BrandOut,
    CategoryOut,
    ProductCreate,
    ProductListResponse,
    ProductOut,
    ProductUpdate,
    RecentReviewOut,
)

router = APIRouter(prefix="/products", tags=["catalog"])
brands_router = APIRouter(prefix="/brands", tags=["catalog"])
categories_router = APIRouter(prefix="/categories", tags=["catalog"])
admin_products_router = APIRouter(
    prefix="/admin/products",
    tags=["admin-catalog"],
    dependencies=[Depends(require_permission("admin"))],
)


async def _supplier_map(db: DBSession, product_ids: list) -> dict:
    """Devuelve {product_id: (supplier_id_str, supplier_name)} usando el ultimo
    proveedor asociado en purchasing.supplier_sku_map (por last_seen_at)."""
    if not product_ids:
        return {}
    rows = (
        await db.execute(
            select(
                SupplierSkuMap.product_id,
                SupplierSkuMap.supplier_id,
                Supplier.name,
                SupplierSkuMap.last_seen_at,
                SupplierSkuMap.created_at,
            )
            .join(Supplier, Supplier.id == SupplierSkuMap.supplier_id)
            .where(SupplierSkuMap.product_id.in_(product_ids))
            .where(Supplier.is_active == True)  # noqa: E712
        )
    ).all()
    from datetime import datetime

    _MIN = datetime.min.replace(tzinfo=UTC)

    def _norm(dt):
        if dt is None:
            return _MIN
        return dt if dt.tzinfo is not None else dt.replace(tzinfo=UTC)

    best: dict = {}
    for pid, sid, sname, last_seen, created in rows:
        key = _norm(last_seen or created)
        cur = best.get(pid)
        if cur is None or key >= cur[0]:
            best[pid] = (key, sid, sname)
    return {pid: (str(v[1]), v[2]) for pid, v in best.items()}


async def _upsert_product_supplier(db: DBSession, product: Product, supplier_id) -> None:
    """Asocia (o reasocia) un producto a un proveedor manualmente.

    Crea/actualiza una fila en supplier_sku_map usando el SKU interno como
    sku_proveedor por defecto, para que la derivacion de proveedor lo detecte.
    """
    from datetime import UTC, datetime

    if supplier_id is None:
        return
    sup = (
        await db.execute(select(Supplier).where(Supplier.id == supplier_id))
    ).scalar_one_or_none()
    if sup is None:
        raise HTTPException(status_code=404, detail="Proveedor no encontrado")
    existing = (
        await db.execute(
            select(SupplierSkuMap)
            .where(SupplierSkuMap.supplier_id == supplier_id)
            .where(SupplierSkuMap.product_id == product.id)
        )
    ).scalar_one_or_none()
    now = datetime.now(UTC)
    if existing is not None:
        existing.last_seen_at = now
    else:
        db.add(
            SupplierSkuMap(
                supplier_id=supplier_id,
                sku_proveedor=product.sku,
                product_id=product.id,
                factor_pack=1,
                last_unit_cost=product.cost,
                last_seen_at=now,
            )
        )


# ----------------------------- Products -------------------------------------
@router.get("", response_model=ProductListResponse)
async def list_products(
    db: DBSession,
    q: str | None = Query(None, description="Búsqueda por nombre/sku"),
    brand_id: uuid.UUID | None = None,
    category_id: uuid.UUID | None = None,
    category_slug: str | None = Query(
        None, description="Filtrar por slug de categoría (resuelve UUID internamente)"
    ),
    species: str | None = Query(
        None, description="Filtrar por especie: perro, gato (incluye mixto automáticamente)"
    ),
    supplier_id: uuid.UUID | None = Query(None, description="Filtrar por proveedor asociado"),
    without_supplier: bool = Query(False, description="Solo productos sin proveedor"),
    is_published: bool | None = None,
    is_featured: bool | None = None,
    page: int = Query(1, ge=1),
    per_page: int = Query(24, ge=1, le=100),
):
    stmt = select(Product).where(Product.deleted_at.is_(None))
    count_stmt = select(func.count(Product.id)).where(Product.deleted_at.is_(None))

    if q:
        # Multi-token AND: each word must appear in name or sku (order-independent)
        tokens = q.split()
        for token in tokens:
            like = f"%{token}%"
            cond = or_(Product.name.ilike(like), Product.sku.ilike(like))
            stmt = stmt.where(cond)
            count_stmt = count_stmt.where(cond)
    if brand_id:
        stmt = stmt.where(Product.brand_id == brand_id)
        count_stmt = count_stmt.where(Product.brand_id == brand_id)
    if category_id:
        stmt = stmt.where(Product.category_id == category_id)
        count_stmt = count_stmt.where(Product.category_id == category_id)
    if category_slug:
        cat = (
            await db.execute(
                select(Category)
                .where(func.lower(Category.slug) == category_slug.lower())
                .where(Category.deleted_at.is_(None))
            )
        ).scalar_one_or_none()
        if cat:
            stmt = stmt.where(Product.category_id == cat.id)
            count_stmt = count_stmt.where(Product.category_id == cat.id)
    if species:
        species_cond = or_(
            Product.attributes["species"].astext == species,
            Product.attributes["species"].astext == "mixto",
        )
        stmt = stmt.where(species_cond)
        count_stmt = count_stmt.where(species_cond)
    if is_published is not None:
        stmt = stmt.where(Product.is_published == is_published)
        count_stmt = count_stmt.where(Product.is_published == is_published)
    if is_featured is not None:
        stmt = stmt.where(Product.is_featured == is_featured)
        count_stmt = count_stmt.where(Product.is_featured == is_featured)

    # Filtros por proveedor (vía supplier_sku_map)
    if supplier_id is not None:
        sub = select(SupplierSkuMap.product_id).where(SupplierSkuMap.supplier_id == supplier_id)
        stmt = stmt.where(Product.id.in_(sub))
        count_stmt = count_stmt.where(Product.id.in_(sub))
    if without_supplier:
        sub_all = select(SupplierSkuMap.product_id).distinct()
        stmt = stmt.where(Product.id.notin_(sub_all))
        count_stmt = count_stmt.where(Product.id.notin_(sub_all))

    total = (await db.execute(count_stmt)).scalar_one()
    stmt = stmt.order_by(Product.created_at.desc()).offset((page - 1) * per_page).limit(per_page)
    rows = (await db.execute(stmt)).scalars().all()

    # Fetch stock totals for all products in one query
    product_ids = [r.id for r in rows]
    stock_map: dict = {}
    if product_ids:
        stock_rows = (
            await db.execute(
                select(Stock.product_id, func.sum(Stock.quantity).label("qty"))
                .where(Stock.product_id.in_(product_ids))
                .group_by(Stock.product_id)
            )
        ).all()
        stock_map = {r.product_id: int(r.qty or 0) for r in stock_rows}

    supplier_map = await _supplier_map(db, product_ids)

    items = []
    for r in rows:
        p_out = ProductOut.model_validate(r)
        p_out.stock_qty = stock_map.get(r.id, 0)
        p_out.in_stock = p_out.stock_qty > 0
        sup = supplier_map.get(r.id)
        if sup:
            p_out.supplier_id = sup[0]
            p_out.supplier_name = sup[1]
        items.append(p_out)

    return ProductListResponse(
        items=items,
        total=total,
        page=page,
        per_page=per_page,
        pages=max(1, math.ceil(total / per_page)),
    )


# ----------------------------- Advanced catalog with facets -----------------
@router.get("/catalog")
async def catalog_products(
    db: DBSession,
    category_slug: str | None = None,
    life_stage: str | None = None,
    size_range: str | None = None,
    health_concerns: str | None = None,
    brand: str | None = None,
    pet_type: str | None = None,
    price_min: float | None = None,
    price_max: float | None = None,
    in_stock: bool | None = None,
    sort: str = Query("relevance", pattern="^(relevance|price_asc|price_desc|name|newest)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(40, ge=1, le=100),
):
    """Advanced catalog endpoint with facet filtering and counts."""
    from sqlalchemy import text as sql_text

    stmt = select(Product).where(
        Product.is_published == True,  # noqa: E712
        Product.deleted_at.is_(None),
    )

    if category_slug:
        cat = (
            await db.execute(
                select(Category).where(func.lower(Category.slug) == category_slug.lower())
            )
        ).scalar_one_or_none()
        if cat:
            stmt = stmt.where(Product.category_id == cat.id)

    if life_stage:
        stages = [s.strip() for s in life_stage.split(",")]
        stmt = stmt.where(Product.life_stage.in_(stages))

    if size_range:
        sizes = [s.strip() for s in size_range.split(",")]
        stmt = stmt.where(Product.size_range.in_(sizes))

    if brand:
        brands_list = [b.strip() for b in brand.split(",")]
        stmt = stmt.where(Product.brand_normalized.in_(brands_list))

    if pet_type:
        stmt = stmt.where(
            or_(
                Product.pet_type == pet_type,
                Product.pet_type == "both",
            )
        )

    if price_min is not None:
        stmt = stmt.where(Product.price >= price_min)
    if price_max is not None:
        stmt = stmt.where(Product.price <= price_max)

    if health_concerns:
        concerns = [c.strip() for c in health_concerns.split(",")]
        stmt = stmt.where(
            sql_text("health_concerns && ARRAY[:concerns]::text[]").bindparams(concerns=concerns)
        )

    if sort == "price_asc":
        stmt = stmt.order_by(Product.price.asc())
    elif sort == "price_desc":
        stmt = stmt.order_by(Product.price.desc())
    elif sort == "name":
        stmt = stmt.order_by(Product.name.asc())
    else:
        stmt = stmt.order_by(Product.created_at.desc())

    count_stmt = select(func.count()).select_from(stmt.subquery())
    total = (await db.execute(count_stmt)).scalar_one()

    stmt = stmt.offset((page - 1) * page_size).limit(page_size)
    rows = (await db.execute(stmt)).scalars().all()

    product_ids = [r.id for r in rows]
    stock_map: dict = {}
    if product_ids:
        stock_rows = (
            await db.execute(
                select(Stock.product_id, func.sum(Stock.quantity).label("qty"))
                .where(Stock.product_id.in_(product_ids))
                .group_by(Stock.product_id)
            )
        ).all()
        stock_map = {r.product_id: int(r.qty or 0) for r in stock_rows}

    items = []
    for r in rows:
        stock_qty = stock_map.get(r.id, 0)
        if in_stock and stock_qty == 0:
            continue
        p_out = ProductOut.model_validate(r)
        p_out.stock_qty = stock_qty
        p_out.in_stock = stock_qty > 0
        items.append(p_out)

    facets: dict = {}
    try:
        facet_stmt = select(Product).where(
            Product.is_published == True,  # noqa: E712
            Product.deleted_at.is_(None),
        )
        facet_rows = (await db.execute(facet_stmt)).scalars().all()

        life_stages: dict = {}
        size_ranges: dict = {}
        brands_count: dict = {}
        pet_types: dict = {}

        for p in facet_rows:
            if p.life_stage and p.life_stage != "all":
                life_stages[p.life_stage] = life_stages.get(p.life_stage, 0) + 1
            if p.size_range and p.size_range != "all":
                size_ranges[p.size_range] = size_ranges.get(p.size_range, 0) + 1
            if p.brand_normalized:
                brands_count[p.brand_normalized] = brands_count.get(p.brand_normalized, 0) + 1
            if p.pet_type and p.pet_type != "both":
                pet_types[p.pet_type] = pet_types.get(p.pet_type, 0) + 1

        facets = {
            "life_stages": life_stages,
            "size_ranges": size_ranges,
            "brands": dict(sorted(brands_count.items(), key=lambda x: -x[1])[:20]),
            "pet_types": pet_types,
        }
    except Exception:
        facets = {}

    return {
        "items": items,
        "total": total,
        "page": page,
        "page_size": page_size,
        "facets": facets,
    }


@router.get("/{product_id}", response_model=ProductOut)
async def get_product(product_id: uuid.UUID, db: DBSession):
    p = (await db.execute(select(Product).where(Product.id == product_id))).scalar_one_or_none()
    if p is None or p.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    p_out = ProductOut.model_validate(p)
    stock_rows = (
        await db.execute(select(func.sum(Stock.quantity)).where(Stock.product_id == p.id))
    ).scalar_one()
    p_out.stock_qty = int(stock_rows or 0)
    p_out.in_stock = p_out.stock_qty > 0
    supplier_map = await _supplier_map(db, [p.id])
    sup = supplier_map.get(p.id)
    if sup:
        p_out.supplier_id = sup[0]
        p_out.supplier_name = sup[1]
    return p_out


@router.get("/by-slug/{slug}", response_model=ProductOut)
async def get_product_by_slug(slug: str, db: DBSession):
    p = (await db.execute(select(Product).where(Product.slug == slug))).scalar_one_or_none()
    if p is None or p.deleted_at is not None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    p_out = ProductOut.model_validate(p)
    stock_total = (
        await db.execute(select(func.sum(Stock.quantity)).where(Stock.product_id == p.id))
    ).scalar_one()
    p_out.stock_qty = int(stock_total or 0)
    p_out.in_stock = p_out.stock_qty > 0

    # Reseñas recientes aprobadas para JSON-LD (max 5)
    review_rows = (
        await db.execute(
            select(ProductReview, Customer.full_name)
            .join(Customer, Customer.id == ProductReview.customer_id, isouter=True)
            .where(ProductReview.product_id == p.id)
            .where(ProductReview.status == "approved")
            .order_by(ProductReview.created_at.desc())
            .limit(5)
        )
    ).all()
    p_out.recent_reviews = [
        RecentReviewOut(
            id=rev.id,
            rating=rev.rating,
            title=rev.title,
            comment=rev.comment,
            reviewer_name=(full_name.split()[0] if full_name else "Cliente"),
            photo_urls=rev.photo_urls or [],
            helpful_count=rev.helpful_count,
            created_at=rev.created_at,
        )
        for rev, full_name in review_rows
    ]
    return p_out


@router.get("/{product_id}/related", response_model=list[ProductOut])
async def related_products(
    product_id: uuid.UUID,
    db: DBSession,
    limit: int = Query(4, ge=1, le=12),
):
    """Productos relacionados (misma categoría, sin el propio producto)."""
    product = (
        await db.execute(
            select(Product).where(Product.id == product_id).where(Product.deleted_at.is_(None))
        )
    ).scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    stmt = (
        select(Product)
        .where(Product.deleted_at.is_(None))
        .where(Product.is_published == True)  # noqa: E712
        .where(Product.id != product_id)
    )
    if product.category_id:
        stmt = stmt.where(Product.category_id == product.category_id)
    stmt = stmt.order_by(Product.is_featured.desc(), Product.created_at.desc()).limit(limit)
    rows = (await db.execute(stmt)).scalars().all()

    product_ids = [r.id for r in rows]
    stock_map: dict = {}
    if product_ids:
        stock_rows = (
            await db.execute(
                select(Stock.product_id, func.sum(Stock.quantity).label("qty"))
                .where(Stock.product_id.in_(product_ids))
                .group_by(Stock.product_id)
            )
        ).all()
        stock_map = {r.product_id: int(r.qty or 0) for r in stock_rows}

    items = []
    for r in rows:
        p_out = ProductOut.model_validate(r)
        p_out.stock_qty = stock_map.get(r.id, 0)
        p_out.in_stock = p_out.stock_qty > 0
        items.append(p_out)
    return items


@router.post(
    "",
    response_model=ProductOut,
    status_code=status.HTTP_201_CREATED,
    dependencies=[Depends(require_permission("catalog:write"))],
)
async def create_product(payload: ProductCreate, db: DBSession):
    data = payload.model_dump()
    supplier_id = data.pop("supplier_id", None)
    slug = slugify(payload.name)
    p = Product(slug=slug, **data)
    db.add(p)
    await db.flush()
    if supplier_id is not None:
        await _upsert_product_supplier(db, p, supplier_id)
    await db.commit()
    await db.refresh(p)
    out = ProductOut.model_validate(p)
    if supplier_id is not None:
        sup = (await _supplier_map(db, [p.id])).get(p.id)
        if sup:
            out.supplier_id = sup[0]
            out.supplier_name = sup[1]
    return out


@router.patch(
    "/{product_id}",
    response_model=ProductOut,
    dependencies=[Depends(require_permission("catalog:write"))],
)
async def update_product(product_id: uuid.UUID, payload: ProductUpdate, db: DBSession):
    p = (await db.execute(select(Product).where(Product.id == product_id))).scalar_one_or_none()
    if p is None:
        raise HTTPException(status_code=404, detail="Producto no encontrado")
    data = payload.model_dump(exclude_unset=True)
    supplier_id = data.pop("supplier_id", None)
    for k, v in data.items():
        setattr(p, k, v)
    # Auto-sync brand_normalized cuando cambia la marca
    if "brand_id" in data:
        if data["brand_id"]:
            brand_obj = (
                await db.execute(select(Brand).where(Brand.id == data["brand_id"]))
            ).scalar_one_or_none()
            if brand_obj and not data.get("brand_normalized"):
                p.brand_normalized = brand_obj.slug.replace("-", "_")
        elif "brand_normalized" not in data:
            p.brand_normalized = None
    if supplier_id is not None:
        await _upsert_product_supplier(db, p, supplier_id)
    await db.commit()
    await db.refresh(p)
    out = ProductOut.model_validate(p)
    sup = (await _supplier_map(db, [p.id])).get(p.id)
    if sup:
        out.supplier_id = sup[0]
        out.supplier_name = sup[1]
    return out


# ─── Admin: Export / Import filtros de catálogo ──────────────────────────────

_FILTER_COLS = [
    ("sku", "SKU", True),
    ("nombre", "Nombre", True),
    ("categoria", "Categoría", True),
    ("marca_display", "Marca (actual)", True),
    ("precio", "Precio", True),
    ("publicado", "Publicado", True),
    ("marca_normalizada", "Marca normalizada", False),
    ("tipo_mascota", "Tipo de mascota", False),
    ("etapa_vida", "Etapa de vida", False),
    ("tamaño_raza", "Tamaño de raza", False),
    ("problemas_salud", "Problemas de salud", False),
]

_DROPDOWNS = {
    "tipo_mascota": '"dog,cat,both,small_pet"',
    "etapa_vida": '"puppy,adult,senior,all"',
    "tamaño_raza": '"mini,small,medium,large,giant,all"',
}


@admin_products_router.get("/export-xlsx")
async def export_products_xlsx(db: DBSession):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.worksheet.datavalidation import DataValidation

    rows = (
        (
            await db.execute(
                select(Product).where(Product.deleted_at.is_(None)).order_by(Product.name)
            )
        )
        .scalars()
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    hdr_fill = PatternFill(start_color="1B5E20", fill_type="solid")
    lock_fill = PatternFill(start_color="E8E8E8", fill_type="solid")
    edit_fill = PatternFill(start_color="FFFDE7", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")

    # Encabezados
    for col_i, (_, label, _locked) in enumerate(_FILTER_COLS, 1):
        cell = ws.cell(row=1, column=col_i, value=label)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Datos
    for row_i, p in enumerate(rows, 2):
        brand_name = p.brand.name if p.brand else ""
        cat_name = p.category.name if p.category else ""
        hc_str = ", ".join(p.health_concerns) if p.health_concerns else ""

        values = [
            p.sku,
            p.name,
            cat_name,
            brand_name,
            float(p.price) if p.price else 0,
            "Sí" if p.is_published else "No",
            p.brand_normalized or "",
            p.pet_type or "",
            p.life_stage or "",
            p.size_range or "",
            hc_str,
        ]
        for col_i, (val, (_key, _label, locked)) in enumerate(
            zip(values, _FILTER_COLS, strict=False), 1
        ):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill = lock_fill if locked else edit_fill
            cell.alignment = Alignment(vertical="center")

    # Data validation (dropdowns) para columnas editables
    n = len(rows) + 1
    for key, formula in _DROPDOWNS.items():
        col_i = next(i for i, (k, _, _) in enumerate(_FILTER_COLS, 1) if k == key)
        dv = DataValidation(type="list", formula1=formula, allow_blank=True, showDropDown=False)
        col_letter = ws.cell(row=1, column=col_i).column_letter
        dv.sqref = f"{col_letter}2:{col_letter}{n}"
        ws.add_data_validation(dv)

    # Anchos de columna
    widths = [18, 45, 22, 22, 10, 10, 22, 16, 14, 16, 30]
    for col_i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=productos_filtros.xlsx"},
    )


@admin_products_router.post("/import-xlsx")
async def import_products_xlsx(file: UploadFile = File(...), db: DBSession = ...):
    from openpyxl import load_workbook

    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Se requiere un archivo .xlsx")

    data = await file.read()
    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="Archivo Excel inválido o dañado") from exc

    ws = wb.active
    updated = skipped = errors = 0
    error_details: list[str] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        sku = str(row[0]).strip()
        brand_normalized = str(row[6]).strip() if row[6] is not None else None
        pet_type = str(row[7]).strip() if row[7] is not None else None
        life_stage = str(row[8]).strip() if row[8] is not None else None
        size_range = str(row[9]).strip() if row[9] is not None else None
        health_raw = str(row[10]).strip() if row[10] is not None else None

        health_concerns = (
            [x.strip() for x in health_raw.split(",") if x.strip()] if health_raw else None
        ) or None

        try:
            result = await db.execute(
                update(Product)
                .where(Product.sku == sku, Product.deleted_at.is_(None))
                .values(
                    brand_normalized=brand_normalized or None,
                    pet_type=pet_type or None,
                    life_stage=life_stage or None,
                    size_range=size_range or None,
                    health_concerns=health_concerns,
                    updated_at=datetime.now(UTC),
                )
            )
            if result.rowcount > 0:
                updated += 1
            else:
                skipped += 1
        except Exception as e:
            errors += 1
            error_details.append(f"SKU {sku}: {e}")

    await db.commit()
    return {
        "ok": errors == 0,
        "updated": updated,
        "skipped": skipped,
        "errors": errors,
        "error_details": error_details[:20],
    }


# ----------------------------- Brands ---------------------------------------
@brands_router.get("", response_model=list[BrandOut])
async def list_brands(db: DBSession):
    rows = (
        (await db.execute(select(Brand).where(Brand.deleted_at.is_(None)).order_by(Brand.name)))
        .scalars()
        .all()
    )
    return rows


# ----------------------------- Categories -----------------------------------
@categories_router.get("", response_model=list[CategoryOut])
async def list_categories(db: DBSession):
    rows = (
        (
            await db.execute(
                select(Category)
                .where(Category.deleted_at.is_(None))
                .order_by(Category.sort_order, Category.name)
            )
        )
        .scalars()
        .all()
    )
    return rows
