"""Export / import masivo de productos en Excel."""

from __future__ import annotations

import io
import uuid
from datetime import UTC, datetime

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select, text

from app.deps import DBSession, require_permission
from app.models.catalog import Brand, Category, Product

catalog_export_router = APIRouter(
    prefix="/catalog",
    tags=["catalog-export"],
    dependencies=[Depends(require_permission("admin"))],
)

# ── Paleta ────────────────────────────────────────────────────────────────────
_TEAL_DARK = "0D4A45"
_TEAL_MID = "187F77"
_TEAL_LIGHT = "E0F2F1"
_HDR_EDITABLE = "FFF9C4"   # amarillo suave — columna editable
_HDR_READONLY = "B2DFDB"   # teal claro — columna de solo lectura
_ROW_ALT = "F0FDF4"
_ROW_WHITE = "FFFFFF"
_LOCK_BG = "FAFAFA"

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Columnas: (header, attr, editable, width)
_COLS = [
    ("ID",                  "id",                  False, 38),
    ("SKU",                 "sku",                 True,  16),
    ("Nombre",              "name",                True,  40),
    ("Categoría",           "category_name",       True,  22),
    ("Marca",               "brand_name",          True,  20),
    ("Descripción Corta",   "short_description",   True,  45),
    ("Descripción",         "description",         True,  60),
    ("Precio Venta",        "price",               True,  16),
    ("Precio Costo",        "cost",                True,  16),
    ("Precio Tachado",      "compare_at_price",    True,  16),
    ("Stock",               "stock",               False, 10),
    ("Activo",              "is_active_label",     True,  10),
    ("Publicado",           "is_published_label",  True,  12),
    ("Destacado",           "is_featured_label",   True,  12),
    ("Tiene Imagen",        "has_image",           False, 14),
    ("URL Imagen Principal","primary_image_url",   True,  50),
    ("Num. Imágenes",       "num_images",          False, 14),
    ("Tipo Mascota",        "pet_type",            True,  14),
    ("Etapa de Vida",       "life_stage",          True,  16),
    ("Tamaño",              "size_range",          True,  14),
    ("Peso (kg/g)",         "peso",                True,  14),
    ("Tags",                "tags_csv",            True,  35),
    ("SEO Título",          "seo_title",           True,  40),
    ("SEO Descripción",     "seo_description",     True,  55),
    ("Creado",              "created_at_label",    False, 20),
]


def _cell_style(ws, row: int, col: int, editable: bool, alt: bool):
    cell = ws.cell(row=row, column=col)
    cell.border = _BORDER
    cell.alignment = Alignment(vertical="top", wrap_text=False)
    if not editable:
        cell.fill = PatternFill("solid", fgColor=_LOCK_BG)
    elif alt:
        cell.fill = PatternFill("solid", fgColor=_ROW_ALT)
    else:
        cell.fill = PatternFill("solid", fgColor=_ROW_WHITE)
    return cell


async def _load_all_products(db: DBSession) -> list[dict]:
    rows = await db.execute(text("""
        SELECT
            p.id, p.sku, p.name, p.slug,
            p.short_description, p.description,
            p.price::float, p.cost::float,
            p.compare_at_price::float,
            p.is_active, p.is_published, p.is_featured,
            p.primary_image_url,
            COALESCE(jsonb_array_length(p.images), 0) AS num_images,
            p.pet_type, p.life_stage, p.size_range,
            p.attributes,
            p.tags,
            p.seo_title, p.seo_description,
            p.created_at,
            c.name AS category_name,
            b.name AS brand_name,
            COALESCE(SUM(s.quantity), 0)::int AS stock
        FROM catalog.products p
        LEFT JOIN catalog.categories c ON c.id = p.category_id
        LEFT JOIN catalog.brands     b ON b.id = p.brand_id
        LEFT JOIN inventory.stock    s ON s.product_id = p.id
        WHERE p.deleted_at IS NULL
        GROUP BY p.id, c.name, b.name
        ORDER BY c.name NULLS LAST, p.name
    """))
    products = []
    for r in rows.mappings():
        attrs = r["attributes"] or {}
        peso = attrs.get("peso") or attrs.get("weight") or attrs.get("Peso") or ""
        tags = r["tags"] or []
        products.append({
            "id": str(r["id"]),
            "sku": r["sku"] or "",
            "name": r["name"] or "",
            "category_name": r["category_name"] or "",
            "brand_name": r["brand_name"] or "",
            "short_description": r["short_description"] or "",
            "description": r["description"] or "",
            "price": float(r["price"] or 0),
            "cost": float(r["cost"] or 0),
            "compare_at_price": float(r["compare_at_price"]) if r["compare_at_price"] else "",
            "stock": int(r["stock"] or 0),
            "is_active_label": "SÍ" if r["is_active"] else "NO",
            "is_published_label": "SÍ" if r["is_published"] else "NO",
            "is_featured_label": "SÍ" if r["is_featured"] else "NO",
            "has_image": "SÍ" if r["primary_image_url"] else "NO",
            "primary_image_url": r["primary_image_url"] or "",
            "num_images": r["num_images"],
            "pet_type": r["pet_type"] or "",
            "life_stage": r["life_stage"] or "",
            "size_range": r["size_range"] or "",
            "peso": str(peso),
            "tags_csv": ", ".join(tags) if isinstance(tags, list) else "",
            "seo_title": r["seo_title"] or "",
            "seo_description": r["seo_description"] or "",
            "created_at_label": r["created_at"].strftime("%Y-%m-%d") if r["created_at"] else "",
        })
    return products


def _build_excel(products: list[dict], categories: list[str], brands: list[str]) -> bytes:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Productos"

    # ── Instrucciones ────────────────────────────────────────────────────────
    ws_info = wb.create_sheet("INSTRUCCIONES", 0)
    ws_info.sheet_view.showGridLines = False
    ws_info["A1"] = "INSTRUCCIONES DE USO"
    ws_info["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws_info["A1"].fill = PatternFill("solid", fgColor=_TEAL_DARK)
    ws_info["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_info.merge_cells("A1:E1")
    ws_info.row_dimensions[1].height = 30

    instrucciones = [
        ("", ""),
        ("EXPORTAR", "Este Excel contiene TODOS tus productos con su información actual."),
        ("EDITAR", "Modifica las columnas EN AMARILLO. Las columnas en gris son solo lectura."),
        ("COLUMNA ID", "NO borres ni modifiques la columna ID — es la clave de importación."),
        ("ACTIVO/PUBLICADO", "Usa SÍ o NO (en mayúsculas) para los campos booleanos."),
        ("CATEGORÍA", "Escribe el nombre exacto de la categoría o déjalo vacío."),
        ("MARCA", "Escribe el nombre exacto de la marca o déjalo vacío."),
        ("TAGS", "Separados por coma. Ej: premium, adulto, pollo"),
        ("PESO", "Escribe el peso con unidad. Ej: 3kg, 500g, 2.5kg"),
        ("IMPORTAR", "Guarda el archivo y súbelo en el admin → Catálogo → Importar Excel."),
        ("", ""),
        ("CAMPOS EDITABLES", "SKU, Nombre, Categoría, Marca, Descripción Corta, Descripción,"),
        ("", "Precio Venta, Precio Costo, Precio Tachado, Activo, Publicado,"),
        ("", "Destacado, URL Imagen Principal, Tipo Mascota, Etapa de Vida,"),
        ("", "Tamaño, Peso, Tags, SEO Título, SEO Descripción"),
    ]
    for i, (label, detail) in enumerate(instrucciones, start=2):
        ws_info.cell(row=i, column=1, value=label).font = Font(bold=True, color=_TEAL_MID)
        ws_info.cell(row=i, column=2, value=detail)
    ws_info.column_dimensions["A"].width = 20
    ws_info.column_dimensions["B"].width = 80

    # ── Sheet de productos ───────────────────────────────────────────────────
    ws = wb["Productos"]
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    # Fila 1: mega-título
    ws.merge_cells(f"A1:{get_column_letter(len(_COLS))}1")
    title_cell = ws["A1"]
    title_cell.value = f"Bigotes y Paticas — Catálogo de Productos ({len(products)} productos)"
    title_cell.font = Font(bold=True, size=13, color="FFFFFF")
    title_cell.fill = PatternFill("solid", fgColor=_TEAL_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Fila 2: encabezados
    for col_i, (header, _, editable, width) in enumerate(_COLS, start=1):
        cell = ws.cell(row=2, column=col_i, value=header)
        cell.font = Font(bold=True, size=10, color="FFFFFF" if not editable else "0D4A45")
        cell.fill = PatternFill("solid", fgColor=_TEAL_MID if not editable else "FFF176")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[2].height = 32

    # Datos
    for row_i, prod in enumerate(products, start=3):
        alt = (row_i % 2 == 1)
        for col_i, (_, attr, editable, _) in enumerate(_COLS, start=1):
            val = prod.get(attr, "")
            cell = _cell_style(ws, row_i, col_i, editable, alt)
            cell.value = val
            if attr in ("price", "cost") and isinstance(val, (int, float)):
                cell.number_format = "#,##0"
            elif attr == "compare_at_price" and isinstance(val, (int, float)) and val:
                cell.number_format = "#,##0"
        ws.row_dimensions[row_i].height = 16

    # Dropdowns para Activo / Publicado / Destacado / PetType
    last_row = 2 + len(products)

    dv_sino = DataValidation(type="list", formula1='"SÍ,NO"', allow_blank=True)
    ws.add_data_validation(dv_sino)

    dv_pet = DataValidation(type="list", formula1='"perro,gato,ambos,"', allow_blank=True)
    ws.add_data_validation(dv_pet)

    # Columnas de SÍ/NO: Activo(12), Publicado(13), Destacado(14)
    for col_i in [12, 13, 14]:
        col_letter = get_column_letter(col_i)
        dv_sino.add(f"{col_letter}3:{col_letter}{last_row}")

    # Pet type col 18
    pet_col = get_column_letter(18)
    dv_pet.add(f"{pet_col}3:{pet_col}{last_row}")

    # Ocultar columna ID (col 1) — la dejamos visible pero angosta para que el
    # usuario no la borre accidentalmente; la marcamos en gris oscuro
    ws.column_dimensions["A"].width = 38

    # Hoja de listas para dropdowns de categoría y marca
    ws_lists = wb.create_sheet("_listas")
    ws_lists.sheet_state = "hidden"
    for i, cat in enumerate(sorted(categories), start=1):
        ws_lists.cell(row=i, column=1, value=cat)
    for i, br in enumerate(sorted(brands), start=1):
        ws_lists.cell(row=i, column=2, value=br)

    n_cats = len(categories)
    n_brands = len(brands)
    if n_cats > 0:
        dv_cat = DataValidation(
            type="list",
            formula1=f"_listas!$A$1:$A${n_cats}",
            allow_blank=True,
        )
        ws.add_data_validation(dv_cat)
        cat_col = get_column_letter(4)
        dv_cat.add(f"{cat_col}3:{cat_col}{last_row}")

    if n_brands > 0:
        dv_brand = DataValidation(
            type="list",
            formula1=f"_listas!$B$1:$B${n_brands}",
            allow_blank=True,
        )
        ws.add_data_validation(dv_brand)
        brand_col = get_column_letter(5)
        dv_brand.add(f"{brand_col}3:{brand_col}{last_row}")

    # Leyenda de colores debajo de los datos
    legend_row = last_row + 2
    ws.cell(row=legend_row, column=1, value="Amarillo = editable").font = Font(size=9, italic=True, color="888888")
    ws.cell(row=legend_row, column=3, value="Gris = solo lectura").font = Font(size=9, italic=True, color="888888")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@catalog_export_router.get("/export-excel")
async def export_products_excel(db: DBSession):
    """Descarga todos los productos en Excel para edición masiva."""
    products = await _load_all_products(db)

    cats = await db.execute(
        select(Category.name).where(Category.is_active == True).where(Category.deleted_at == None)  # noqa: E711,E712
    )
    brands = await db.execute(
        select(Brand.name).where(Brand.is_active == True).where(Brand.deleted_at == None)  # noqa: E711,E712
    )
    cat_names = [r[0] for r in cats.all() if r[0]]
    brand_names = [r[0] for r in brands.all() if r[0]]

    data = _build_excel(products, cat_names, brand_names)
    filename = f"BigotesyPaticas_Productos_{datetime.now(UTC).strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        io.BytesIO(data),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── Importar ─────────────────────────────────────────────────────────────────

_COL_MAP = {
    "ID": "id",
    "SKU": "sku",
    "Nombre": "name",
    "Categoría": "category_name",
    "Marca": "brand_name",
    "Descripción Corta": "short_description",
    "Descripción": "description",
    "Precio Venta": "price",
    "Precio Costo": "cost",
    "Precio Tachado": "compare_at_price",
    "Activo": "is_active_label",
    "Publicado": "is_published_label",
    "Destacado": "is_featured_label",
    "URL Imagen Principal": "primary_image_url",
    "Tipo Mascota": "pet_type",
    "Etapa de Vida": "life_stage",
    "Tamaño": "size_range",
    "Peso (kg/g)": "peso",
    "Tags": "tags_csv",
    "SEO Título": "seo_title",
    "SEO Descripción": "seo_description",
}


def _slugify_simple(text_val: str) -> str:
    import unicodedata, re
    t = unicodedata.normalize("NFKD", text_val).encode("ascii", "ignore").decode("ascii")
    t = re.sub(r"[^\w\s-]", "", t.lower())
    return re.sub(r"[\s_-]+", "-", t).strip("-") or "item"


async def _get_or_create_category(db, cat_name: str, cat_map: dict) -> uuid.UUID | None:
    """Devuelve el ID de la categoría, creándola si no existe."""
    if not cat_name:
        return None
    key = cat_name.strip().lower()
    if key in cat_map:
        return cat_map[key]
    # Crear nueva
    new_id = uuid.uuid4()
    slug = _slugify_simple(cat_name)
    existing = (await db.execute(
        text("SELECT COUNT(*) FROM catalog.categories WHERE slug = :s"), {"s": slug}
    )).scalar()
    if existing:
        slug = f"{slug}-{str(new_id)[:4]}"
    await db.execute(text("""
        INSERT INTO catalog.categories (id, name, slug, is_active, sort_order, created_at, updated_at)
        VALUES (:id::uuid, :name, :slug, true, 0, NOW(), NOW())
    """), {"id": str(new_id), "name": cat_name.strip(), "slug": slug})
    cat_map[key] = new_id
    return new_id


async def _get_or_create_brand(db, brand_name: str, brand_map: dict) -> uuid.UUID | None:
    """Devuelve el ID de la marca, creándola si no existe."""
    if not brand_name:
        return None
    key = brand_name.strip().lower()
    if key in brand_map:
        return brand_map[key]
    new_id = uuid.uuid4()
    slug = _slugify_simple(brand_name)
    existing = (await db.execute(
        text("SELECT COUNT(*) FROM catalog.brands WHERE slug = :s"), {"s": slug}
    )).scalar()
    if existing:
        slug = f"{slug}-{str(new_id)[:4]}"
    await db.execute(text("""
        INSERT INTO catalog.brands (id, name, slug, is_active, created_at, updated_at)
        VALUES (:id::uuid, :name, :slug, true, NOW(), NOW())
    """), {"id": str(new_id), "name": brand_name.strip(), "slug": slug})
    brand_map[key] = new_id
    return new_id


@catalog_export_router.post("/import-excel")
async def import_products_excel(
    db: DBSession,
    file: UploadFile = File(...),
):
    """Importa el Excel: actualiza existentes, crea nuevos, crea categorías/marcas nuevas."""
    import json as _json

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Archivo Excel inválido")

    if "Productos" not in wb.sheetnames:
        raise HTTPException(400, "El archivo no tiene hoja 'Productos'")

    ws = wb["Productos"]

    # Encabezados en fila 1 o 2 (tolerante)
    headers: dict[int, str] = {}
    for cell in ws[2]:
        if cell.value and str(cell.value).strip() in _COL_MAP:
            headers[cell.column] = _COL_MAP[str(cell.value).strip()]
    if not headers:
        for cell in ws[1]:
            if cell.value and str(cell.value).strip() in _COL_MAP:
                headers[cell.column] = _COL_MAP[str(cell.value).strip()]

    # Cargar categorías y marcas actuales (case-insensitive)
    cats_rows = await db.execute(
        select(Category.id, Category.name).where(Category.deleted_at == None)  # noqa: E711
    )
    brands_rows = await db.execute(
        select(Brand.id, Brand.name).where(Brand.deleted_at == None)  # noqa: E711
    )
    cat_map: dict[str, uuid.UUID] = {r.name.strip().lower(): r.id for r in cats_rows.all()}
    brand_map: dict[str, uuid.UUID] = {r.name.strip().lower(): r.id for r in brands_rows.all()}

    # SKUs existentes para detectar duplicados al crear
    sku_rows = await db.execute(
        text("SELECT sku FROM catalog.products WHERE deleted_at IS NULL")
    )
    existing_skus = {r[0] for r in sku_rows.all()}

    updated = created = 0
    errors: list[str] = []

    def _str(v) -> str | None:
        return str(v).strip() if v not in (None, "") else None

    def _float(v) -> float | None:
        try:
            return float(v) if v not in (None, "") else None
        except (ValueError, TypeError):
            return None

    def _bool(v, default: bool = False) -> bool:
        if v is None:
            return default
        return str(v).strip().upper() in ("SÍ", "SI", "S", "YES", "TRUE", "1")

    data_start = 3 if "id" in headers.values() else 2
    for row in ws.iter_rows(min_row=data_start, values_only=True):
        if not any(v for v in row):
            continue

        row_data: dict[str, object] = {
            field: row[col_i - 1]
            for col_i, field in headers.items()
            if col_i <= len(row)
        }

        product_id_raw = row_data.get("id")
        sku_raw = _str(row_data.get("sku"))
        name_raw = _str(row_data.get("name"))

        # ── Resolver categoría y marca (crear si no existen) ──────────────────
        cat_name_raw  = _str(row_data.get("category_name"))
        brand_name_raw = _str(row_data.get("brand_name"))
        cat_id   = await _get_or_create_category(db, cat_name_raw or "", cat_map)
        brand_id = await _get_or_create_brand(db, brand_name_raw or "", brand_map)

        peso_val = _str(row_data.get("peso"))
        tags_raw = row_data.get("tags_csv")
        tag_list: list[str] = (
            [t.strip() for t in str(tags_raw).split(",") if t.strip()]
            if tags_raw not in (None, "") else []
        )

        # ── CREAR producto nuevo (sin ID o ID no en BD) ───────────────────────
        is_new = False
        try:
            pid = uuid.UUID(str(product_id_raw).strip()) if product_id_raw else None
        except (ValueError, AttributeError):
            pid = None

        if pid is None:
            # Fila sin ID → crear producto nuevo
            if not name_raw:
                errors.append("Fila sin ID ni Nombre — omitida")
                continue
            if not sku_raw:
                import re as _re
                sku_raw = _re.sub(r"[^\w-]", "-", name_raw.lower())[:64]
            if sku_raw in existing_skus:
                errors.append(f"SKU duplicado al crear: '{sku_raw}' — omitido")
                continue
            pid = uuid.uuid4()
            slug_base = _slugify_simple(name_raw)
            slug_check = (await db.execute(
                text("SELECT COUNT(*) FROM catalog.products WHERE slug LIKE :s"),
                {"s": f"{slug_base}%"}
            )).scalar()
            slug = slug_base if not slug_check else f"{slug_base}-{str(pid)[:4]}"
            await db.execute(text("""
                INSERT INTO catalog.products
                  (id, sku, name, slug, category_id, brand_id,
                   short_description, description,
                   price, cost, compare_at_price,
                   is_active, is_published, is_featured,
                   primary_image_url, pet_type, life_stage, size_range,
                   seo_title, seo_description,
                   tags, attributes,
                   images, created_at, updated_at)
                VALUES
                  (:id::uuid, :sku, :name, :slug, :cat_id::uuid, :brand_id::uuid,
                   :short_desc, :desc,
                   :price, :cost, :cap,
                   :is_active, :is_published, :is_featured,
                   :img, :pet_type, :life_stage, :size_range,
                   :seo_t, :seo_d,
                   :tags::jsonb, :attrs::jsonb,
                   '[]'::jsonb, NOW(), NOW())
            """), {
                "id": str(pid), "sku": sku_raw, "name": name_raw, "slug": slug,
                "cat_id": str(cat_id) if cat_id else None,
                "brand_id": str(brand_id) if brand_id else None,
                "short_desc": _str(row_data.get("short_description")),
                "desc": _str(row_data.get("description")),
                "price": _float(row_data.get("price")) or 0,
                "cost": _float(row_data.get("cost")) or 0,
                "cap": _float(row_data.get("compare_at_price")),
                "is_active": _bool(row_data.get("is_active_label"), True),
                "is_published": _bool(row_data.get("is_published_label"), False),
                "is_featured": _bool(row_data.get("is_featured_label"), False),
                "img": _str(row_data.get("primary_image_url")),
                "pet_type": _str(row_data.get("pet_type")),
                "life_stage": _str(row_data.get("life_stage")),
                "size_range": _str(row_data.get("size_range")),
                "seo_t": _str(row_data.get("seo_title")),
                "seo_d": _str(row_data.get("seo_description")),
                "tags": _json.dumps(tag_list, ensure_ascii=False),
                "attrs": _json.dumps({"peso": peso_val} if peso_val else {}, ensure_ascii=False),
            })
            existing_skus.add(sku_raw)
            created += 1
            continue

        # ── ACTUALIZAR producto existente ─────────────────────────────────────
        params: dict[str, object] = {
            "pid": str(pid),
            "cat_id": str(cat_id) if cat_id else None,
            "brand_id": str(brand_id) if brand_id else None,
            "name": name_raw,
            "short_desc": _str(row_data.get("short_description")),
            "desc": _str(row_data.get("description")),
            "price": _float(row_data.get("price")),
            "cost": _float(row_data.get("cost")),
            "cap": _float(row_data.get("compare_at_price")),
            "is_active": _bool(row_data.get("is_active_label"), True) if row_data.get("is_active_label") is not None else None,
            "is_published": _bool(row_data.get("is_published_label"), False) if row_data.get("is_published_label") is not None else None,
            "is_featured": _bool(row_data.get("is_featured_label"), False) if row_data.get("is_featured_label") is not None else None,
            "img": _str(row_data.get("primary_image_url")),
            "pet_type": _str(row_data.get("pet_type")),
            "life_stage": _str(row_data.get("life_stage")),
            "size_range": _str(row_data.get("size_range")),
            "seo_t": _str(row_data.get("seo_title")),
            "seo_d": _str(row_data.get("seo_description")),
            "tags": _json.dumps(tag_list, ensure_ascii=False),
        }
        if sku_raw:
            params["sku"] = sku_raw

        await db.execute(text("""
            UPDATE catalog.products SET
                name              = COALESCE(:name, name),
                category_id       = :cat_id::uuid,
                brand_id          = :brand_id::uuid,
                short_description = :short_desc,
                description       = :desc,
                price             = COALESCE(:price, price),
                cost              = COALESCE(:cost, cost),
                compare_at_price  = :cap,
                is_active         = COALESCE(:is_active, is_active),
                is_published      = COALESCE(:is_published, is_published),
                is_featured       = COALESCE(:is_featured, is_featured),
                primary_image_url = COALESCE(:img, primary_image_url),
                pet_type          = :pet_type,
                life_stage        = :life_stage,
                size_range        = :size_range,
                seo_title         = :seo_t,
                seo_description   = :seo_d,
                tags              = :tags::jsonb,
                updated_at        = NOW()
            WHERE id = :pid::uuid AND deleted_at IS NULL
        """), params)

        if peso_val:
            await db.execute(text("""
                UPDATE catalog.products
                SET attributes = attributes || jsonb_build_object('peso', :peso::text),
                    updated_at = NOW()
                WHERE id = :pid::uuid
            """), {"peso": peso_val, "pid": str(pid)})

        updated += 1

    await db.commit()

    return {
        "updated": updated,
        "created": created,
        "errors": errors[:50],
        "total_errors": len(errors),
        "message": f"{updated} actualizados, {created} creados"
                   + (f", {len(errors)} advertencias" if errors else ""),
    }
